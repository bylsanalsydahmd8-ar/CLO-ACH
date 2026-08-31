from flask import (
    render_template,
    url_for,
    flash,
    redirect,
    request,
    send_file
)

from io import BytesIO
from datetime import datetime

from flask_login import (
    login_user,
    logout_user,
    login_required,
    current_user
)

from werkzeug.security import (
    generate_password_hash,
    check_password_hash
)

from sqlalchemy import func

import openpyxl
import re

# =========================================================
# REPORTLAB - PDF
# =========================================================

from reportlab.lib import colors

from reportlab.lib.enums import (
    TA_CENTER,
    TA_LEFT
)

from reportlab.lib.pagesizes import A4

from reportlab.lib.styles import (
    getSampleStyleSheet,
    ParagraphStyle
)

from reportlab.lib.units import mm

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
    KeepTogether
)

from reportlab.graphics.shapes import (
    Drawing,
    String
)

from reportlab.graphics.charts.barcharts import (
    VerticalBarChart
)

# =========================================================
# FLASK APPLICATION
# =========================================================

from student import (
    app,
    db,
    login_manager
)

# =========================================================
# MODELS
# =========================================================

from student.models import (
    User,
    Course,
    AcademicTerm,
    CLO,
    Assessment,
    Question,
    Student,
    Enrollment,
    StudentQuestionResult
)

# =========================================================
# FORMS
# =========================================================

from student.forms import (
    RegistrationForm,
    LoginForm,
    CourseForm,
    CLOForm,
    AssessmentForm,
    QuestionForm,
    StudentForm,
    StudentExcelUploadForm,
    StudentMarksExcelUploadForm
)
# =========================================================
# USER LOADER
# =========================================================

@login_manager.user_loader
def load_user(user_id):

    return User.query.get(int(user_id))


# =========================================================
# HOME
# =========================================================
@app.route("/")
@login_required
def home():

    courses = Course.query.order_by(
        Course.id.desc()
    ).limit(3).all()

    return render_template(
        "home.html",
        courses=courses
    )


# =========================================================
# REGISTER
# =========================================================

@app.route(
    "/register",
    methods=["GET", "POST"]
)
def register():

    if current_user.is_authenticated:

        return redirect(
            url_for("home")
        )

    form = RegistrationForm()

    if form.validate_on_submit():

        hashed_password = generate_password_hash(
            form.password.data
        )

        user = User(

            fname=form.fname.data.strip(),

            lname=form.lname.data.strip(),

            username=form.username.data.strip(),

            email=form.email.data.strip().lower(),

            role=form.role.data,

            password=hashed_password

        )

        db.session.add(user)

        try:

            db.session.commit()

        except Exception as error:

            db.session.rollback()

            print("REGISTER ERROR:")
            print(type(error).__name__)
            print(error)

            flash(
                "Unable to create account.",
                "danger"
            )

            return render_template(
                "register.html",
                title="Register",
                form=form
            )

        flash(
            "Your account has been created successfully!",
            "success"
        )

        return redirect(
            url_for("login")
        )

    return render_template(
        "register.html",
        title="Register",
        form=form
    )


# =========================================================
# LOGIN
# =========================================================

@app.route(
    "/login",
    methods=["GET", "POST"]
)
def login():

    if current_user.is_authenticated:

        return redirect(
            url_for("home")
        )

    form = LoginForm()

    if form.validate_on_submit():

        user = User.query.filter_by(
            email=form.email.data.strip().lower()
        ).first()

        if user is None:

            flash(
                "Invalid email or password.",
                "danger"
            )

            return redirect(
                url_for("login")
            )

        if not check_password_hash(
            user.password,
            form.password.data
        ):

            flash(
                "Invalid email or password.",
                "danger"
            )

            return redirect(
                url_for("login")
            )

        login_user(
            user,
            remember=form.remember.data
        )

        flash(
            f"Welcome back, {user.fname}!",
            "success"
        )

        return redirect(
            url_for("home")
        )

    return render_template(
        "login.html",
        form=form,
        title="Login"
    )


# =========================================================
# LOGOUT
# =========================================================

@app.route("/logout")
@login_required
def logout():

    logout_user()

    flash(
        "You have been logged out successfully.",
        "success"
    )

    return redirect(
        url_for("login")
    )


# =========================================================
# ALL COURSES
# =========================================================

@app.route("/courses")
@login_required
def courses():

    courses = Course.query.filter_by(
        instructor_id=current_user.id
    ).order_by(
        Course.course_code
    ).all()

    return render_template(
        "courses.html",
        courses=courses
    )


# =========================================================
# CREATE COURSE
# =========================================================

@app.route(
    "/courses/create",
    methods=["GET", "POST"]
)
@login_required
def create_course():

    form = CourseForm()

    current_year = 2026

    form.academic_year.choices = [

        (
            f"{year}-{year + 1}",
            f"{year}-{year + 1}"
        )

        for year in range(
            2020,
            current_year + 1
        )

    ]

    if form.validate_on_submit():

        academic_term = AcademicTerm.query.filter_by(

            academic_year=form.academic_year.data,

            semester=form.semester.data

        ).first()

        if academic_term is None:

            academic_term = AcademicTerm(

                academic_year=form.academic_year.data,

                semester=form.semester.data

            )

            db.session.add(
                academic_term
            )

            db.session.flush()

        course = Course(

            course_code=form.course_code.data.strip(),

            course_name=form.course_name.data.strip(),

            description=(
                form.description.data.strip()
                if form.description.data
                else None
            ),

            instructor_id=current_user.id,

            academic_term_id=academic_term.id

        )

        db.session.add(course)

        try:

            db.session.commit()

        except Exception as error:

            db.session.rollback()

            print("CREATE COURSE ERROR:")
            print(type(error).__name__)
            print(error)

            flash(
                "Unable to create course.",
                "danger"
            )

            return render_template(
                "create_course.html",
                form=form
            )

        flash(
            "Course created successfully!",
            "success"
        )

        return redirect(
            url_for("courses")
        )

    return render_template(
        "create_course.html",
        form=form
    )


# =========================================================
# VIEW COURSE
# =========================================================

@app.route(
    "/courses/<int:course_id>"
)
@login_required
def view_course(course_id):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    return render_template(
        "view_course.html",
        course=course
    )


# =========================================================
# EDIT COURSE
# =========================================================

@app.route(
    "/courses/<int:course_id>/edit",
    methods=["GET", "POST"]
)
@login_required
def edit_course(course_id):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    form = CourseForm()

    current_year = 2026

    form.academic_year.choices = [

        (
            f"{year}-{year + 1}",
            f"{year}-{year + 1}"
        )

        for year in range(
            2020,
            current_year + 1
        )

    ]

    if form.validate_on_submit():

        academic_term = AcademicTerm.query.filter_by(

            academic_year=form.academic_year.data,

            semester=form.semester.data

        ).first()

        if academic_term is None:

            academic_term = AcademicTerm(

                academic_year=form.academic_year.data,

                semester=form.semester.data

            )

            db.session.add(
                academic_term
            )

            db.session.flush()

        course.course_code = (
            form.course_code.data.strip()
        )

        course.course_name = (
            form.course_name.data.strip()
        )

        course.description = (

            form.description.data.strip()

            if form.description.data

            else None

        )

        course.academic_term_id = (
            academic_term.id
        )

        try:

            db.session.commit()

        except Exception as error:

            db.session.rollback()

            print("EDIT COURSE ERROR:")
            print(type(error).__name__)
            print(error)

            flash(
                "Unable to update course.",
                "danger"
            )

            return render_template(
                "edit_course.html",
                form=form,
                course=course
            )

        flash(
            "Course updated successfully!",
            "success"
        )

        return redirect(
            url_for(
                "view_course",
                course_id=course.id
            )
        )

    if request.method == "GET":

        form.course_code.data = (
            course.course_code
        )

        form.course_name.data = (
            course.course_name
        )

        form.description.data = (
            course.description
        )

        if course.academic_term:

            form.academic_year.data = (
                course.academic_term.academic_year
            )

            form.semester.data = (
                course.academic_term.semester
            )

    return render_template(
        "edit_course.html",
        form=form,
        course=course
    )


# =========================================================
# DELETE COURSE
# =========================================================

@app.route(
    "/courses/<int:course_id>/delete",
    methods=["POST"]
)
@login_required
def delete_course(course_id):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    course_code = course.course_code

    db.session.delete(course)

    try:

        db.session.commit()

    except Exception as error:

        db.session.rollback()

        print("DELETE COURSE ERROR:")
        print(type(error).__name__)
        print(error)

        flash(
            "Unable to delete this course.",
            "danger"
        )

        return redirect(
            url_for(
                "view_course",
                course_id=course_id
            )
        )

    flash(
        f"Course {course_code} deleted successfully.",
        "success"
    )

    return redirect(
        url_for("courses")
    )


# =========================================================
# =========================================================
# STUDENT MANAGEMENT
# =========================================================
# =========================================================


# =========================================================
# STUDENT SELECT COURSE
# =========================================================

@app.route("/students")
@login_required
def students():

    courses = Course.query.filter_by(

        instructor_id=current_user.id

    ).order_by(

        Course.course_code

    ).all()

    return render_template(
        "students.html",
        courses=courses
    )


# =========================================================
# COURSE STUDENTS
# =========================================================

@app.route(
    "/courses/<int:course_id>/students"
)
@login_required
def course_students(course_id):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    enrollments = (

        Enrollment.query

        .filter_by(
            course_id=course.id
        )

        .join(Student)

        .order_by(
            Student.student_number
        )

        .all()

    )

    return render_template(

        "course_students.html",

        course=course,

        enrollments=enrollments

    )


# =========================================================
# ADD STUDENT
# =========================================================

@app.route(
    "/courses/<int:course_id>/students/add",
    methods=["GET", "POST"]
)
@login_required
def add_student(course_id):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    form = StudentForm()

    if form.validate_on_submit():

        student_number = (
            form.student_number.data.strip()
        )

        full_name = (
            form.student_name.data.strip()
        )

        student = Student.query.filter_by(

            student_number=student_number

        ).first()

        if student:

            existing_enrollment = (
                Enrollment.query.filter_by(

                    student_id=student.id,

                    course_id=course.id

                ).first()
            )

            if existing_enrollment:

                flash(
                    "This student is already enrolled in this course.",
                    "warning"
                )

                return redirect(
                    url_for(
                        "course_students",
                        course_id=course.id
                    )
                )

            name_parts = full_name.split()

            student.fname = name_parts[0]

            student.lname = (

                " ".join(name_parts[1:])

                if len(name_parts) > 1

                else ""

            )

        else:

            name_parts = full_name.split()

            student = Student(

                student_number=student_number,

                fname=name_parts[0],

                lname=(

                    " ".join(name_parts[1:])

                    if len(name_parts) > 1

                    else ""

                )

            )

            db.session.add(student)

            db.session.flush()

        enrollment = Enrollment(

            student_id=student.id,

            course_id=course.id

        )

        db.session.add(enrollment)

        try:

            db.session.commit()

        except Exception as error:

            db.session.rollback()

            print("ADD STUDENT ERROR:")
            print(type(error).__name__)
            print(error)

            flash(
                "Unable to add student.",
                "danger"
            )

            return render_template(
                "add_student.html",
                form=form,
                course=course
            )

        flash(
            f"Student {full_name} added successfully.",
            "success"
        )

        return redirect(
            url_for(
                "course_students",
                course_id=course.id
            )
        )

    return render_template(
        "add_student.html",
        form=form,
        course=course
    )


# =========================================================
# EDIT STUDENT
# =========================================================

@app.route(
    "/courses/<int:course_id>/students/<int:student_id>/edit",
    methods=["GET", "POST"]
)
@login_required
def edit_student(
    course_id,
    student_id
):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    enrollment = Enrollment.query.filter_by(

        course_id=course.id,

        student_id=student_id

    ).first_or_404()

    student = enrollment.student

    form = StudentForm()

    if request.method == "GET":

        form.student_number.data = (
            student.student_number
        )

        form.student_name.data = (
            f"{student.fname} {student.lname}"
        ).strip()

    if form.validate_on_submit():

        new_student_number = (
            form.student_number.data.strip()
        )

        new_full_name = (
            form.student_name.data.strip()
        )

        existing_student = Student.query.filter(

            Student.student_number ==
            new_student_number,

            Student.id != student.id

        ).first()

        if existing_student:

            flash(
                "This university ID already belongs to another student.",
                "danger"
            )

            return render_template(
                "edit_student.html",
                form=form,
                course=course,
                student=student
            )

        name_parts = new_full_name.split()

        student.fname = name_parts[0]

        student.lname = (

            " ".join(name_parts[1:])

            if len(name_parts) > 1

            else ""

        )

        student.student_number = new_student_number

        try:

            db.session.commit()

        except Exception as error:

            db.session.rollback()

            print("EDIT STUDENT ERROR:")
            print(type(error).__name__)
            print(error)

            flash(
                "Unable to update student.",
                "danger"
            )

            return render_template(
                "edit_student.html",
                form=form,
                course=course,
                student=student
            )

        flash(
            "Student updated successfully.",
            "success"
        )

        return redirect(
            url_for(
                "course_students",
                course_id=course.id
            )
        )

    return render_template(
        "edit_student.html",
        form=form,
        course=course,
        student=student
    )


# =========================================================
# REMOVE STUDENT
# =========================================================

@app.route(
    "/courses/<int:course_id>/students/<int:student_id>/remove",
    methods=["POST"]
)
@login_required
def remove_student_from_course(
    course_id,
    student_id
):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    enrollment = Enrollment.query.filter_by(

        course_id=course.id,

        student_id=student_id

    ).first_or_404()

    student_name = (

        f"{enrollment.student.fname} "
        f"{enrollment.student.lname}"

    ).strip()

    db.session.delete(enrollment)

    try:

        db.session.commit()

    except Exception as error:

        db.session.rollback()

        print("REMOVE STUDENT ERROR:")
        print(type(error).__name__)
        print(error)

        flash(
            "Unable to remove student from this course.",
            "danger"
        )

        return redirect(
            url_for(
                "course_students",
                course_id=course.id
            )
        )

    flash(
        f"{student_name} removed from the course.",
        "success"
    )

    return redirect(
        url_for(
            "course_students",
            course_id=course.id
        )
    )


# =========================================================
# IMPORT STUDENTS FROM EXCEL
# =========================================================

@app.route(
    "/courses/<int:course_id>/students/import",
    methods=["GET", "POST"]
)
@login_required
def import_students_excel(course_id):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    form = StudentExcelUploadForm()

    if form.validate_on_submit():

        try:

            workbook = openpyxl.load_workbook(

                form.file.data,

                data_only=True

            )

            worksheet = workbook.active

            headers = [

                str(cell.value)
                .strip()
                .lower()

                if cell.value is not None

                else ""

                for cell in worksheet[1]

            ]

            student_number_index = None
            student_name_index = None

            for index, header in enumerate(headers):

                normalized = (

                    header
                    .replace("_", " ")
                    .replace("-", " ")
                    .strip()

                )

                if normalized in (

                    "university id",
                    "student id",
                    "student number",
                    "student no",
                    "student no.",
                    "id",
                    "الرقم الجامعي",
                    "رقم الطالب"

                ):

                    student_number_index = index

                if normalized in (

                    "student name",
                    "name",
                    "full name",
                    "student",
                    "الاسم",
                    "اسم الطالب"

                ):

                    student_name_index = index

            if (

                student_number_index is None

                or

                student_name_index is None

            ):

                flash(

                    "Excel file must contain "
                    "'University ID' and "
                    "'Student Name' columns.",

                    "danger"

                )

                return render_template(

                    "import_students.html",

                    form=form,

                    course=course

                )

            added_count = 0
            existing_count = 0
            skipped_count = 0

            errors = []

            for row_number, row in enumerate(

                worksheet.iter_rows(
                    min_row=2,
                    values_only=True
                ),

                start=2

            ):

                if not row:
                    continue

                student_number = row[
                    student_number_index
                ]

                student_name = row[
                    student_name_index
                ]

                if (

                    student_number is None

                    and

                    student_name is None

                ):

                    continue

                if student_number is None:

                    skipped_count += 1

                    errors.append(

                        f"Row {row_number}: "
                        "University ID is missing."

                    )

                    continue

                student_number = str(
                    student_number
                ).strip()

                if student_number.endswith(".0"):

                    student_number = (
                        student_number[:-2]
                    )

                if not student_number:

                    skipped_count += 1

                    errors.append(

                        f"Row {row_number}: "
                        "University ID is empty."

                    )

                    continue

                if student_name is None:

                    skipped_count += 1

                    errors.append(

                        f"Row {row_number}: "
                        "Student name is missing."

                    )

                    continue

                student_name = str(
                    student_name
                ).strip()

                if not student_name:

                    skipped_count += 1

                    errors.append(

                        f"Row {row_number}: "
                        "Student name is empty."

                    )

                    continue

                name_parts = student_name.split()

                student = Student.query.filter_by(

                    student_number=student_number

                ).first()

                if student is None:

                    student = Student(

                        student_number=student_number,

                        fname=name_parts[0],

                        lname=(

                            " ".join(name_parts[1:])

                            if len(name_parts) > 1

                            else ""

                        )

                    )

                    db.session.add(student)

                    db.session.flush()

                else:

                    existing_count += 1

                    student.fname = name_parts[0]

                    student.lname = (

                        " ".join(name_parts[1:])

                        if len(name_parts) > 1

                        else ""

                    )

                enrollment = Enrollment.query.filter_by(

                    student_id=student.id,

                    course_id=course.id

                ).first()

                if enrollment:
                    continue

                db.session.add(

                    Enrollment(

                        student_id=student.id,

                        course_id=course.id

                    )

                )

                added_count += 1

            db.session.commit()

            flash(

                f"Excel import completed. "
                f"{added_count} students added. "
                f"{existing_count} existing students processed. "
                f"{skipped_count} rows skipped.",

                "success"

            )

            for message in errors[:10]:

                flash(
                    message,
                    "warning"
                )

            if len(errors) > 10:

                flash(

                    f"{len(errors) - 10} more errors "
                    "were not displayed.",

                    "warning"

                )

        except Exception as error:

            db.session.rollback()

            print("STUDENT EXCEL IMPORT ERROR:")
            print(type(error).__name__)
            print(error)

            flash(

                "Unable to import students from Excel. "
                "Please check the file format.",

                "danger"

            )

            return render_template(

                "import_students.html",

                form=form,

                course=course

            )

        return redirect(

            url_for(

                "course_students",

                course_id=course.id

            )

        )

    return render_template(

        "import_students.html",

        form=form,

        course=course

    )


# =========================================================
# =========================================================
# CLO MANAGEMENT
# =========================================================
# =========================================================

# =========================================================
# ALL CLOs
# =========================================================

@app.route("/clos")
@login_required
def clos():

    courses = Course.query.filter_by(

        instructor_id=current_user.id

    ).order_by(

        Course.course_code

    ).all()

    selected_course_id = request.args.get(
        "course_id",
        type=int
    )

    selected_course = None
    selected_clos = []

    if selected_course_id:

        selected_course = Course.query.filter_by(

            id=selected_course_id,

            instructor_id=current_user.id

        ).first()

        if selected_course is None:

            flash(
                "Invalid course selection.",
                "danger"
            )

            return redirect(
                url_for("clos")
            )

        selected_clos = CLO.query.filter_by(

            course_id=selected_course.id

        ).order_by(

            CLO.clo_code

        ).all()

    return render_template(

        "clos.html",

        courses=courses,

        selected_course=selected_course,

        selected_clos=selected_clos

    )


# =========================================================
# COURSE CLOs
# =========================================================

@app.route(
    "/courses/<int:course_id>/clos"
)
@login_required
def course_clos(course_id):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    clos = CLO.query.filter_by(

        course_id=course.id

    ).order_by(

        CLO.clo_code

    ).all()

    return render_template(

        "course_clos.html",

        course=course,

        clos=clos

    )


# =========================================================
# CREATE CLO
# =========================================================

@app.route(
    "/courses/<int:course_id>/clos/create",
    methods=["GET", "POST"]
)
@login_required
def create_clo(course_id):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    clo_count = CLO.query.filter_by(

        course_id=course.id

    ).count()

    if clo_count >= 6:

        flash(
            "This course cannot have more than 6 CLOs.",
            "danger"
        )

        return redirect(

            url_for(
                "course_clos",
                course_id=course.id
            )

        )

    form = CLOForm()

    if form.validate_on_submit():

        clo_code = (
            form.clo_code.data
            .strip()
            .upper()
        )

        description = (

            form.description.data.strip()

            if form.description.data

            else ""

        )

        existing_clo = CLO.query.filter_by(

            course_id=course.id,

            clo_code=clo_code

        ).first()

        if existing_clo:

            flash(
                f"{clo_code} already exists in this course.",
                "danger"
            )

            return render_template(

                "create_clo.html",

                form=form,

                course=course

            )

        clo = CLO(

            clo_code=clo_code,

            description=description,

            category=form.category.data,

            course_id=course.id

        )

        db.session.add(clo)

        try:

            db.session.commit()

        except Exception as error:

            db.session.rollback()

            print("CREATE CLO ERROR:")
            print(type(error).__name__)
            print(error)

            flash(
                "Unable to create CLO.",
                "danger"
            )

            return render_template(

                "create_clo.html",

                form=form,

                course=course

            )

        flash(
            f"{clo_code} created successfully.",
            "success"
        )

        return redirect(

            url_for(
                "course_clos",
                course_id=course.id
            )

        )

    return render_template(

        "create_clo.html",

        form=form,

        course=course

    )


# =========================================================
# EDIT CLO
# =========================================================

@app.route(
    "/courses/<int:course_id>/clos/<int:clo_id>/edit",
    methods=["GET", "POST"]
)
@login_required
def edit_clo(
    course_id,
    clo_id
):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    clo = CLO.query.filter_by(

        id=clo_id,

        course_id=course.id

    ).first_or_404()

    form = CLOForm(
        obj=clo
    )

    if form.validate_on_submit():

        clo_code = (
            form.clo_code.data
            .strip()
            .upper()
        )

        existing_clo = CLO.query.filter(

            CLO.course_id == course.id,

            CLO.clo_code == clo_code,

            CLO.id != clo.id

        ).first()

        if existing_clo:

            flash(
                f"{clo_code} already exists in this course.",
                "danger"
            )

            return render_template(

                "edit_clo.html",

                form=form,

                course=course,

                clo=clo

            )

        clo.clo_code = clo_code

        clo.description = (

            form.description.data.strip()

            if form.description.data

            else ""

        )

        clo.category = form.category.data

        try:

            db.session.commit()

        except Exception as error:

            db.session.rollback()

            print("EDIT CLO ERROR:")
            print(type(error).__name__)
            print(error)

            flash(
                "Unable to update CLO.",
                "danger"
            )

            return render_template(

                "edit_clo.html",

                form=form,

                course=course,

                clo=clo

            )

        flash(
            f"{clo_code} updated successfully.",
            "success"
        )

        return redirect(

            url_for(
                "course_clos",
                course_id=course.id
            )

        )

    return render_template(

        "edit_clo.html",

        form=form,

        course=course,

        clo=clo

    )


# =========================================================
# DELETE CLO
# =========================================================

@app.route(
    "/courses/<int:course_id>/clos/<int:clo_id>/delete",
    methods=["POST"]
)
@login_required
def delete_clo(
    course_id,
    clo_id
):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    clo = CLO.query.filter_by(

        id=clo_id,

        course_id=course.id

    ).first_or_404()

    code = clo.clo_code

    db.session.delete(clo)

    try:

        db.session.commit()

    except Exception as error:

        db.session.rollback()

        print("DELETE CLO ERROR:")
        print(type(error).__name__)
        print(error)

        flash(
            "Unable to delete this CLO.",
            "danger"
        )

        return redirect(

            url_for(
                "course_clos",
                course_id=course.id
            )

        )

    flash(
        f"{code} deleted successfully.",
        "success"
    )

    return redirect(

        url_for(
            "course_clos",
            course_id=course.id
        )

    )


# =========================================================
# =========================================================
# ASSESSMENT MANAGEMENT
# =========================================================
# =========================================================


# =========================================================
# ASSESSMENTS PAGE
# =========================================================

@app.route("/assessments")
@login_required
def assessments():

    courses = Course.query.filter_by(

        instructor_id=current_user.id

    ).order_by(

        Course.course_code

    ).all()

    selected_course_id = request.args.get(

        "course_id",
        type=int

    )

    selected_course = None
    selected_assessments = []

    if selected_course_id:

        selected_course = Course.query.filter_by(

            id=selected_course_id,

            instructor_id=current_user.id

        ).first()

        if selected_course is None:

            flash(
                "Invalid course selection.",
                "danger"
            )

            return redirect(
                url_for("assessments")
            )

        selected_assessments = Assessment.query.filter_by(

            course_id=selected_course.id

        ).order_by(

            Assessment.id

        ).all()

    return render_template(

        "assessments.html",

        courses=courses,

        selected_course=selected_course,

        selected_assessments=selected_assessments

    )


# =========================================================
# COURSE ASSESSMENTS
# =========================================================

@app.route(
    "/courses/<int:course_id>/assessments"
)
@login_required
def course_assessments(course_id):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    assessments = Assessment.query.filter_by(

        course_id=course.id

    ).order_by(

        Assessment.id

    ).all()

    return render_template(

        "course_assessments.html",

        course=course,

        assessments=assessments

    )


# =========================================================
# CREATE ASSESSMENT
# =========================================================

@app.route(
    "/courses/<int:course_id>/assessments/create",
    methods=["GET", "POST"]
)
@login_required
def create_assessment(course_id):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    form = AssessmentForm()

    if form.validate_on_submit():

        try:

            max_mark = float(
                form.max_mark.data
            )

            weight = float(
                form.weight.data
            )

            if max_mark <= 0:

                flash(
                    "Assessment maximum mark must be greater than zero.",
                    "danger"
                )

                return render_template(

                    "create_assessment.html",

                    form=form,

                    course=course

                )

            if weight < 0 or weight > 100:

                flash(
                    "Assessment weight must be between 0 and 100.",
                    "danger"
                )

                return render_template(

                    "create_assessment.html",

                    form=form,

                    course=course

                )

            current_weight = db.session.query(

                func.coalesce(

                    func.sum(
                        Assessment.weight
                    ),

                    0

                )

            ).filter(

                Assessment.course_id == course.id

            ).scalar()

            if current_weight + weight > 100:

                flash(

                    f"Cannot create this assessment. "
                    f"Only {100 - current_weight}% is remaining.",

                    "danger"

                )

                return render_template(

                    "create_assessment.html",

                    form=form,

                    course=course

                )

            assessment = Assessment(

                name=form.name.data.strip(),

                description=(

                    form.description.data.strip()

                    if form.description.data

                    else None

                ),

                max_mark=max_mark,

                weight=weight,

                assessment_type=form.assessment_type.data,

                course_id=course.id

            )

            db.session.add(
                assessment
            )

            db.session.commit()

            flash(
                "Assessment created successfully.",
                "success"
            )

            return redirect(

                url_for(

                    "view_assessment",

                    course_id=course.id,

                    assessment_id=assessment.id

                )

            )

        except Exception as error:

            db.session.rollback()

            print("CREATE ASSESSMENT ERROR:")
            print(type(error).__name__)
            print(error)

            flash(
                "Unable to create assessment.",
                "danger"
            )

    return render_template(

        "create_assessment.html",

        form=form,

        course=course

    )


# =========================================================
# VIEW ASSESSMENT
# =========================================================

@app.route(
    "/courses/<int:course_id>/assessments/<int:assessment_id>"
)
@login_required
def view_assessment(
    course_id,
    assessment_id
):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    assessment = Assessment.query.filter_by(

        id=assessment_id,

        course_id=course.id

    ).first_or_404()

    questions = Question.query.filter_by(

        assessment_id=assessment.id

    ).order_by(

        Question.id

    ).all()

    question_total = sum(

        question.max_mark

        for question in questions

    )

    remaining_mark = (

        assessment.max_mark
        - question_total

    )

    return render_template(

        "view_assessment.html",

        course=course,

        assessment=assessment,

        questions=questions,

        question_total=question_total,

        remaining_mark=remaining_mark

    )


# =========================================================
# EDIT ASSESSMENT
# =========================================================

@app.route(
    "/courses/<int:course_id>/assessments/"
    "<int:assessment_id>/edit",
    methods=["GET", "POST"]
)
@login_required
def edit_assessment(
    course_id,
    assessment_id
):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    assessment = Assessment.query.filter_by(

        id=assessment_id,

        course_id=course.id

    ).first_or_404()

    form = AssessmentForm(
        obj=assessment
    )

    total_question_marks = db.session.query(

        func.coalesce(

            func.sum(
                Question.max_mark
            ),

            0

        )

    ).filter(

        Question.assessment_id == assessment.id

    ).scalar()

    total_question_marks = float(
        total_question_marks or 0
    )

    if form.validate_on_submit():

        new_max_mark = float(
            form.max_mark.data
        )

        weight = float(
            form.weight.data
        )

        if new_max_mark <= 0:

            flash(
                "Assessment maximum mark must be greater than zero.",
                "danger"
            )

            return render_template(

                "edit_assessment.html",

                form=form,

                course=course,

                assessment=assessment,

                total_question_marks=total_question_marks

            )

        if total_question_marks > new_max_mark:

            flash(

                f"Cannot reduce the assessment mark to "
                f"{new_max_mark}. "
                f"The current questions total "
                f"{total_question_marks} marks.",

                "danger"

            )

            return render_template(

                "edit_assessment.html",

                form=form,

                course=course,

                assessment=assessment,

                total_question_marks=total_question_marks

            )

        if weight < 0 or weight > 100:

            flash(
                "Assessment weight must be between 0 and 100.",
                "danger"
            )

            return render_template(

                "edit_assessment.html",

                form=form,

                course=course,

                assessment=assessment,

                total_question_marks=total_question_marks

            )

        assessment.name = (
            form.name.data.strip()
        )

        assessment.description = (

            form.description.data.strip()

            if form.description.data

            else None

        )

        assessment.max_mark = new_max_mark

        assessment.weight = weight

        assessment.assessment_type = (
            form.assessment_type.data
        )

        try:

            db.session.commit()

        except Exception as error:

            db.session.rollback()

            print("EDIT ASSESSMENT ERROR:")
            print(type(error).__name__)
            print(error)

            flash(
                "Unable to update assessment.",
                "danger"
            )

            return render_template(

                "edit_assessment.html",

                form=form,

                course=course,

                assessment=assessment,

                total_question_marks=total_question_marks

            )

        flash(
            "Assessment updated successfully.",
            "success"
        )

        return redirect(

            url_for(

                "view_assessment",

                course_id=course.id,

                assessment_id=assessment.id

            )

        )

    return render_template(

        "edit_assessment.html",

        form=form,

        course=course,

        assessment=assessment,

        total_question_marks=total_question_marks

    )


# =========================================================
# DELETE ASSESSMENT
# =========================================================

@app.route(
    "/courses/<int:course_id>/assessments/"
    "<int:assessment_id>/delete",
    methods=["POST"]
)
@login_required
def delete_assessment(
    course_id,
    assessment_id
):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    assessment = Assessment.query.filter_by(

        id=assessment_id,

        course_id=course.id

    ).first_or_404()

    name = assessment.name

    db.session.delete(assessment)

    try:

        db.session.commit()

    except Exception as error:

        db.session.rollback()

        print("DELETE ASSESSMENT ERROR:")
        print(type(error).__name__)
        print(error)

        flash(
            "Unable to delete assessment.",
            "danger"
        )

        return redirect(

            url_for(

                "course_assessments",

                course_id=course.id

            )

        )

    flash(

        f"Assessment '{name}' deleted successfully.",

        "success"

    )

    return redirect(

        url_for(

            "course_assessments",

            course_id=course.id

        )

    )


# =========================================================
# =========================================================
# QUESTION MANAGEMENT
# =========================================================
# =========================================================


# =========================================================
# CREATE QUESTION
# =========================================================

@app.route(
    "/courses/<int:course_id>/assessments/"
    "<int:assessment_id>/questions/create",
    methods=["GET", "POST"]
)
@login_required
def create_question(
    course_id,
    assessment_id
):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    assessment = Assessment.query.filter_by(

        id=assessment_id,

        course_id=course.id

    ).first_or_404()

    form = QuestionForm()

    if form.validate_on_submit():

        question_mark = float(
            form.max_mark.data
        )

        current_total = db.session.query(

            func.coalesce(

                func.sum(
                    Question.max_mark
                ),

                0

            )

        ).filter(

            Question.assessment_id == assessment.id

        ).scalar()

        new_total = (

            current_total
            + question_mark

        )

        if new_total > assessment.max_mark:

            flash(

                f"Question cannot be added. "
                f"Remaining marks are "
                f"{assessment.max_mark - current_total}.",

                "danger"

            )

            return render_template(

                "create_question.html",

                form=form,

                course=course,

                assessment=assessment

            )

        question = Question(

            question_text=(
                form.question_text.data.strip()
            ),

            max_mark=question_mark,

            assessment_id=assessment.id

        )

        db.session.add(question)

        try:

            db.session.commit()

        except Exception as error:

            db.session.rollback()

            print("CREATE QUESTION ERROR:")
            print(type(error).__name__)
            print(error)

            flash(
                "Unable to create question.",
                "danger"
            )

            return render_template(

                "create_question.html",

                form=form,

                course=course,

                assessment=assessment

            )

        flash(
            "Question created successfully.",
            "success"
        )

        return redirect(

            url_for(

                "view_assessment",

                course_id=course.id,

                assessment_id=assessment.id

            )

        )

    return render_template(

        "create_question.html",

        form=form,

        course=course,

        assessment=assessment

    )


# =========================================================
# EDIT QUESTION
# =========================================================

@app.route(
    "/courses/<int:course_id>/assessments/"
    "<int:assessment_id>/questions/"
    "<int:question_id>/edit",
    methods=["GET", "POST"]
)
@login_required
def edit_question(
    course_id,
    assessment_id,
    question_id
):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    assessment = Assessment.query.filter_by(

        id=assessment_id,

        course_id=course.id

    ).first_or_404()

    question = Question.query.filter_by(

        id=question_id,

        assessment_id=assessment.id

    ).first_or_404()

    form = QuestionForm(
        obj=question
    )

    if form.validate_on_submit():

        new_mark = float(
            form.max_mark.data
        )

        other_questions_total = db.session.query(

            func.coalesce(

                func.sum(
                    Question.max_mark
                ),

                0

            )

        ).filter(

            Question.assessment_id == assessment.id,

            Question.id != question.id

        ).scalar()

        if (

            other_questions_total
            + new_mark
            > assessment.max_mark

        ):

            flash(

                f"Question cannot be updated. "
                f"Assessment maximum is "
                f"{assessment.max_mark}.",

                "danger"

            )

            return render_template(

                "edit_question.html",

                form=form,

                course=course,

                assessment=assessment,

                question=question

            )

        question.question_text = (
            form.question_text.data.strip()
        )

        question.max_mark = new_mark

        try:

            db.session.commit()

        except Exception as error:

            db.session.rollback()

            print("EDIT QUESTION ERROR:")
            print(type(error).__name__)
            print(error)

            flash(
                "Unable to update question.",
                "danger"
            )

            return render_template(

                "edit_question.html",

                form=form,

                course=course,

                assessment=assessment,

                question=question

            )

        flash(
            "Question updated successfully.",
            "success"
        )

        return redirect(

            url_for(

                "view_assessment",

                course_id=course.id,

                assessment_id=assessment.id

            )

        )

    return render_template(

        "edit_question.html",

        form=form,

        course=course,

        assessment=assessment,

        question=question

    )


# =========================================================
# DELETE QUESTION
# =========================================================

@app.route(
    "/courses/<int:course_id>/assessments/"
    "<int:assessment_id>/questions/"
    "<int:question_id>/delete",
    methods=["POST"]
)
@login_required
def delete_question(
    course_id,
    assessment_id,
    question_id
):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    assessment = Assessment.query.filter_by(

        id=assessment_id,

        course_id=course.id

    ).first_or_404()

    question = Question.query.filter_by(

        id=question_id,

        assessment_id=assessment.id

    ).first_or_404()

    db.session.delete(question)

    try:

        db.session.commit()

    except Exception as error:

        db.session.rollback()

        print("DELETE QUESTION ERROR:")
        print(type(error).__name__)
        print(error)

        flash(
            "Unable to delete question.",
            "danger"
        )

        return redirect(

            url_for(

                "view_assessment",

                course_id=course.id,

                assessment_id=assessment.id

            )

        )

    flash(
        "Question deleted successfully.",
        "success"
    )

    return redirect(

        url_for(

            "view_assessment",

            course_id=course.id,

            assessment_id=assessment.id

        )

    )


# =========================================================
# =========================================================
# QUESTION - CLO MATRIX
# =========================================================
# =========================================================


# =========================================================
# QUESTION CLO MATRIX
# =========================================================

@app.route(
    "/courses/<int:course_id>/assessments/"
    "<int:assessment_id>/matrix"
)
@login_required
def question_clo_matrix(
    course_id,
    assessment_id
):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    assessment = Assessment.query.filter_by(

        id=assessment_id,

        course_id=course.id

    ).first_or_404()

    clos = CLO.query.filter_by(

        course_id=course.id

    ).order_by(

        CLO.clo_code

    ).all()

    questions = Question.query.filter_by(

        assessment_id=assessment.id

    ).order_by(

        Question.id

    ).all()

    return render_template(

        "question_clo_matrix.html",

        course=course,

        assessment=assessment,

        clos=clos,

        questions=questions

    )


# =========================================================
# TOGGLE QUESTION CLO
# =========================================================

@app.route(
    "/courses/<int:course_id>/assessments/"
    "<int:assessment_id>/questions/"
    "<int:question_id>/clo/"
    "<int:clo_id>/toggle",
    methods=["POST"]
)
@login_required
def toggle_question_clo(
    course_id,
    assessment_id,
    question_id,
    clo_id
):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    assessment = Assessment.query.filter_by(

        id=assessment_id,

        course_id=course.id

    ).first_or_404()

    question = Question.query.filter_by(

        id=question_id,

        assessment_id=assessment.id

    ).first_or_404()

    clo = CLO.query.filter_by(

        id=clo_id,

        course_id=course.id

    ).first_or_404()

    if clo in question.clos:

        question.clos.remove(clo)

        action = "removed"

    else:

        question.clos.append(clo)

        action = "added"

    try:

        db.session.commit()

    except Exception as error:

        db.session.rollback()

        print("QUESTION-CLO ERROR:")
        print(type(error).__name__)
        print(error)

        flash(
            "Unable to update question CLO mapping.",
            "danger"
        )

        return redirect(

            url_for(

                "question_clo_matrix",

                course_id=course.id,

                assessment_id=assessment.id

            )

        )

    if action == "added":

        flash(

            f"{question.question_text} mapped to "
            f"{clo.clo_code}.",

            "success"

        )

    else:

        flash(

            f"{question.question_text} unmapped from "
            f"{clo.clo_code}.",

            "success"

        )

    return redirect(

        url_for(

            "question_clo_matrix",

            course_id=course.id,

            assessment_id=assessment.id

        )

    )


# =========================================================
# EXCEL QUESTION IMPORT
# =========================================================

@app.route(
    "/courses/<int:course_id>/assessments/"
    "<int:assessment_id>/questions/import",
    methods=["POST"]
)
@login_required
def import_questions_excel(
    course_id,
    assessment_id
):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    assessment = Assessment.query.filter_by(

        id=assessment_id,

        course_id=course.id

    ).first_or_404()

    file = request.files.get(
        "excel_file"
    )

    if not file or not file.filename:

        flash(
            "Please select an Excel file.",
            "danger"
        )

        return redirect(

            url_for(

                "view_assessment",

                course_id=course.id,

                assessment_id=assessment.id

            )

        )

    if not file.filename.lower().endswith(
        (".xlsx", ".xlsm")
    ):

        flash(
            "Only Excel .xlsx and .xlsm files are supported.",
            "danger"
        )

        return redirect(

            url_for(

                "view_assessment",

                course_id=course.id,

                assessment_id=assessment.id

            )

        )

    try:

        workbook = openpyxl.load_workbook(
            file,
            data_only=True
        )

        worksheet = workbook.active

        headers = [

            str(cell.value)
            .strip()
            .lower()

            if cell.value is not None

            else ""

            for cell in worksheet[1]

        ]

        question_index = None
        mark_index = None

        for index, header in enumerate(headers):

            if header in (

                "question",
                "question name",
                "question text",
                "name",
                "السؤال",
                "اسم السؤال"

            ):

                question_index = index

            if header in (

                "mark",
                "marks",
                "max mark",
                "score",
                "علامة",
                "العلامة"

            ):

                mark_index = index

        if (

            question_index is None
            or
            mark_index is None

        ):

            flash(

                "Excel file must contain "
                "'Question' and 'Mark' columns.",

                "danger"

            )

            return redirect(

                url_for(

                    "view_assessment",

                    course_id=course.id,

                    assessment_id=assessment.id

                )

            )

        imported_questions = []
        imported_total = 0

        for row in worksheet.iter_rows(

            min_row=2,

            values_only=True

        ):

            if not row:
                continue

            question_text = row[
                question_index
            ]

            mark = row[
                mark_index
            ]

            if (

                question_text is None
                and
                mark is None

            ):
                continue

            if question_text is None:

                raise ValueError(
                    "A question name is missing."
                )

            if mark is None:

                raise ValueError(

                    f"Mark is missing for question "
                    f"'{question_text}'."

                )

            question_text = str(
                question_text
            ).strip()

            try:

                question_mark = float(
                    mark
                )

            except (
                TypeError,
                ValueError
            ):

                raise ValueError(

                    f"Invalid mark for question "
                    f"'{question_text}'."

                )

            if question_mark <= 0:

                raise ValueError(

                    f"Question '{question_text}' "
                    "must have a mark greater than zero."

                )

            imported_total += question_mark

            imported_questions.append(

                (
                    question_text,
                    question_mark
                )

            )

        current_total = db.session.query(

            func.coalesce(

                func.sum(
                    Question.max_mark
                ),

                0

            )

        ).filter(

            Question.assessment_id == assessment.id

        ).scalar()

        if (

            current_total
            + imported_total
            > assessment.max_mark

        ):

            flash(

                f"Cannot import questions. "
                f"Only "
                f"{assessment.max_mark - current_total} "
                f"marks are available.",

                "danger"

            )

            return redirect(

                url_for(

                    "view_assessment",

                    course_id=course.id,

                    assessment_id=assessment.id

                )

            )

        for (

            question_text,
            question_mark

        ) in imported_questions:

            db.session.add(

                Question(

                    question_text=question_text,

                    max_mark=question_mark,

                    assessment_id=assessment.id

                )

            )

        db.session.commit()

        flash(

            f"{len(imported_questions)} questions "
            "imported successfully.",

            "success"

        )

    except ValueError as error:

        db.session.rollback()

        flash(
            str(error),
            "danger"
        )

    except Exception as error:

        db.session.rollback()

        print("QUESTION EXCEL IMPORT ERROR:")
        print(type(error).__name__)
        print(error)

        flash(

            "Unable to import Excel file. "
            "Please check the file format.",

            "danger"

        )

    return redirect(

        url_for(

            "view_assessment",

            course_id=course.id,

            assessment_id=assessment.id

        )

    )


# =========================================================
# =========================================================
# STUDENT MARKS
# =========================================================
# =========================================================


# =========================================================
# EXTRACT QUESTION NUMBER
# =========================================================

def extract_question_number(header):

    if header is None:
        return None

    text = str(
        header
    ).strip()

    if not text:
        return None

    normalized = re.sub(
        r"\s+",
        " ",
        text.lower()
    ).strip()

    # =====================================================
    # Q1
    # Q 1
    # Q.1
    # Q. 1
    # Q_1
    # Q-1
    # Q1/2.00
    # Q. 1 /2.00
    # Q 1 / 2.00
    # =====================================================

    match = re.match(

        r"^\s*q\s*[\._\-]?\s*(\d+)",

        normalized,

        re.IGNORECASE

    )

    if match:

        return int(
            match.group(1)
        )

    # =====================================================
    # Question 1
    # Question_1
    # Question-1
    # Question.1
    # Question 1 / 2.00
    # =====================================================

    match = re.match(

        r"^\s*question\s*[\._\-]?\s*(\d+)",

        normalized,

        re.IGNORECASE

    )

    if match:

        return int(
            match.group(1)
        )

    # =====================================================
    # السؤال 1
    # السؤال_1
    # السؤال-1
    # السؤال.1
    # السؤال 1 / 2.00
    # =====================================================

    match = re.match(

        r"^\s*السؤال\s*[\._\-]?\s*(\d+)",

        normalized

    )

    if match:

        return int(
            match.group(1)
        )

    # =====================================================
    # NUMBER ONLY
    # =====================================================

    match = re.match(

        r"^\s*(\d+)\s*$",

        normalized

    )

    if match:

        return int(
            match.group(1)
        )

    return None


# =========================================================
# IMPORT STUDENT MARKS FROM EXCEL
# =========================================================

@app.route(
    "/courses/<int:course_id>/assessments/"
    "<int:assessment_id>/results/import",
    methods=["GET", "POST"]
)
@login_required
def import_student_marks_excel(
    course_id,
    assessment_id
):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    assessment = Assessment.query.filter_by(

        id=assessment_id,

        course_id=course.id

    ).first_or_404()

    questions = Question.query.filter_by(

        assessment_id=assessment.id

    ).order_by(

        Question.id

    ).all()

    if not questions:

        flash(

            "This assessment does not contain any questions yet.",

            "warning"

        )

        return redirect(

            url_for(

                "view_assessment",

                course_id=course.id,

                assessment_id=assessment.id

            )

        )

    form = StudentMarksExcelUploadForm()

    if form.validate_on_submit():

        try:

            workbook = openpyxl.load_workbook(

                form.file.data,

                data_only=True

            )

            worksheet = workbook.active

            headers = [

                str(cell.value).strip()

                if cell.value is not None

                else ""

                for cell in worksheet[1]

            ]

            # =================================================
            # FIND UNIVERSITY ID
            # =================================================

            student_number_index = None

            for index, header in enumerate(headers):

                normalized = (

                    header
                    .lower()
                    .replace("_", " ")
                    .replace("-", " ")
                    .strip()

                )

                if normalized in (

                    "university id",
                    "student id",
                    "student number",
                    "student no",
                    "student no.",
                    "id",
                    "الرقم الجامعي",
                    "رقم الطالب"

                ):

                    student_number_index = index
                    break

            if student_number_index is None:

                flash(

                    "Excel file must contain "
                    "'University ID' column.",

                    "danger"

                )

                return render_template(

                    "import_student_marks.html",

                    form=form,

                    course=course,

                    assessment=assessment,

                    questions=questions

                )

            # =================================================
            # QUESTION MAP
            # =================================================

            question_map = {}

            for number, question in enumerate(

                questions,

                start=1

            ):

                question_map[
                    f"q{number}"
                ] = question

            # =================================================
            # QUESTION COLUMNS
            # =================================================

            excel_question_columns = {}

            duplicate_headers = []

            ignored_question_headers = []

            for index, header in enumerate(headers):

                question_number = (
                    extract_question_number(
                        header
                    )
                )

                if question_number is None:

                    continue

                question_key = (
                    f"q{question_number}"
                )

                if question_key not in question_map:

                    ignored_question_headers.append(
                        header
                    )

                    continue

                if question_key in excel_question_columns:

                    duplicate_headers.append(
                        header
                    )

                    continue

                excel_question_columns[
                    question_key
                ] = index

            # =================================================
            # NO QUESTION COLUMNS
            # =================================================

            if not excel_question_columns:

                flash(

                    "No question columns were found. "
                    "Accepted formats include: "
                    "Q1, Q2, "
                    "Question 1, Question_1, "
                    "Question-1, "
                    "السؤال 1, السؤال 2, "
                    "Q. 1 /2.00, Q1/2.00.",

                    "danger"

                )

                return render_template(

                    "import_student_marks.html",

                    form=form,

                    course=course,

                    assessment=assessment,

                    questions=questions

                )

            # =================================================
            # MISSING QUESTIONS
            # =================================================

            missing_questions = []

            for question_key in question_map:

                if question_key not in (
                    excel_question_columns
                ):

                    missing_questions.append(
                        question_key.upper()
                    )

            added_count = 0
            updated_count = 0
            skipped_count = 0

            errors = []

            # =================================================
            # PROCESS EXCEL ROWS
            # =================================================

            for row_number, row in enumerate(

                worksheet.iter_rows(
                    min_row=2,
                    values_only=True
                ),

                start=2

            ):

                if not row:
                    continue

                student_number = row[
                    student_number_index
                ]

                if student_number is None:

                    skipped_count += 1

                    errors.append(

                        f"Row {row_number}: "
                        "University ID is missing."

                    )

                    continue

                student_number = str(
                    student_number
                ).strip()

                if student_number.endswith(".0"):

                    student_number = (
                        student_number[:-2]
                    )

                if not student_number:

                    skipped_count += 1

                    errors.append(

                        f"Row {row_number}: "
                        "University ID is empty."

                    )

                    continue

                # =================================================
                # FIND STUDENT
                # =================================================

                student = Student.query.filter_by(

                    student_number=student_number

                ).first()

                if student is None:

                    skipped_count += 1

                    errors.append(

                        f"Row {row_number}: "
                        f"Student {student_number} "
                        "does not exist."

                    )

                    continue

                # =================================================
                # FIND ENROLLMENT
                # =================================================

                enrollment = Enrollment.query.filter_by(

                    student_id=student.id,

                    course_id=course.id

                ).first()

                if enrollment is None:

                    skipped_count += 1

                    errors.append(

                        f"Row {row_number}: "
                        f"Student {student_number} "
                        "is not enrolled in this course."

                    )

                    continue

                # =================================================
                # PROCESS QUESTIONS
                # =================================================

                for (

                    question_key,
                    column_index

                ) in excel_question_columns.items():

                    question = question_map[
                        question_key
                    ]

                    mark_value = row[
                        column_index
                    ]

                    if mark_value is None:
                        continue

                    if (

                        isinstance(
                            mark_value,
                            str
                        )

                        and

                        not mark_value.strip()

                    ):

                        continue

                    try:

                        mark = float(

                            str(
                                mark_value
                            ).strip()

                        )

                    except (

                        TypeError,
                        ValueError

                    ):

                        errors.append(

                            f"Row {row_number}: "
                            f"Invalid mark "
                            f"'{mark_value}' "
                            f"for "
                            f"{question_key.upper()}."

                        )

                        continue

                    # =================================================
                    # NEGATIVE
                    # =================================================

                    if mark < 0:

                        errors.append(

                            f"Row {row_number}: "
                            f"{question_key.upper()} "
                            "cannot be negative."

                        )

                        continue

                    # =================================================
                    # EXCEEDS MAXIMUM
                    # =================================================

                    if mark > question.max_mark:

                        errors.append(

                            f"Row {row_number}: "
                            f"{question_key.upper()} "
                            f"mark {mark} "
                            f"exceeds maximum "
                            f"{question.max_mark}."

                        )

                        continue

                    # =================================================
                    # FIND RESULT
                    # =================================================

                    result = (

                        StudentQuestionResult.query
                        .filter_by(

                            enrollment_id=enrollment.id,

                            question_id=question.id

                        )
                        .first()

                    )

                    # =================================================
                    # UPDATE
                    # =================================================

                    if result:

                        result.mark = mark

                        updated_count += 1

                    # =================================================
                    # CREATE
                    # =================================================

                    else:

                        result = StudentQuestionResult(

                            enrollment_id=enrollment.id,

                            question_id=question.id,

                            mark=mark

                        )

                        db.session.add(
                            result
                        )

                        added_count += 1

            # =================================================
            # COMMIT
            # =================================================

            db.session.commit()

            flash(

                f"Marks import completed. "
                f"{added_count} marks added. "
                f"{updated_count} marks updated. "
                f"{skipped_count} student rows skipped.",

                "success"

            )

            if missing_questions:

                flash(

                    "These question columns were not found: "
                    +
                    ", ".join(
                        missing_questions
                    ),

                    "warning"

                )

            if ignored_question_headers:

                print(
                    "Ignored question headers:"
                )

                for header in ignored_question_headers:

                    print(
                        repr(header)
                    )

            if duplicate_headers:

                print(
                    "Duplicate question headers:"
                )

                for header in duplicate_headers:

                    print(
                        repr(header)
                    )

            for error_message in errors[:10]:

                flash(
                    error_message,
                    "warning"
                )

            if len(errors) > 10:

                flash(

                    f"{len(errors) - 10} more "
                    "errors were not displayed.",

                    "warning"

                )

            return redirect(

                url_for(

                    "assessment_student_marks",

                    course_id=course.id,

                    assessment_id=assessment.id

                )

            )

        except Exception as error:

            db.session.rollback()

            print("====================================")
            print("STUDENT MARKS IMPORT ERROR")
            print("====================================")
            print(type(error).__name__)
            print(error)
            print("====================================")

            flash(

                "Unable to import student marks. "
                "Please check the Excel file.",

                "danger"

            )

            return render_template(

                "import_student_marks.html",

                form=form,

                course=course,

                assessment=assessment,

                questions=questions

            )

    return render_template(

        "import_student_marks.html",

        form=form,

        course=course,

        assessment=assessment,

        questions=questions

    )


# =========================================================
# VIEW STUDENT MARKS
# =========================================================

@app.route(

    "/courses/<int:course_id>/assessments/"
    "<int:assessment_id>/results"

)
@login_required
def assessment_student_marks(

    course_id,
    assessment_id

):

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    assessment = Assessment.query.filter_by(

        id=assessment_id,

        course_id=course.id

    ).first_or_404()

    questions = Question.query.filter_by(

        assessment_id=assessment.id

    ).order_by(

        Question.id

    ).all()

    enrollments = (

        Enrollment.query

        .filter_by(
            course_id=course.id
        )

        .join(Student)

        .order_by(
            Student.student_number
        )

        .all()

    )

    student_marks = []

    for enrollment in enrollments:

        marks = {}

        total_mark = 0.0

        for question in questions:

            result = (

                StudentQuestionResult.query

                .filter_by(

                    enrollment_id=enrollment.id,

                    question_id=question.id

                )

                .first()

            )

            if result:

                mark = float(
                    result.mark
                )

                marks[
                    question.id
                ] = mark

                total_mark += mark

            else:

                marks[
                    question.id
                ] = None

        if assessment.max_mark > 0:

            percentage = (

                total_mark
                /
                assessment.max_mark

            ) * 100

        else:

            percentage = 0

        student_marks.append({

            "enrollment": enrollment,

            "marks": marks,

            "total": total_mark,

            "percentage": percentage

        })

    return render_template(

        "assessment_student_marks.html",

        course=course,

        assessment=assessment,

        questions=questions,

        student_marks=student_marks

    )


# =========================================================
# UPDATE STUDENT QUESTION MARK
# =========================================================
#
# POST
#
# URL example:
#
# /courses/1/assessments/1/results/5/question/2/update
#
# student_id = 5
# question_id = 2
#
# Validation:
#
# 0 <= mark <= question.max_mark
#
# =========================================================

@app.route(

    "/courses/<int:course_id>/assessments/"
    "<int:assessment_id>/results/"
    "<int:student_id>/question/"
    "<int:question_id>/update",

    methods=["POST"]

)
@login_required
def update_student_question_mark(

    course_id,
    assessment_id,
    student_id,
    question_id

):

    # =====================================================
    # COURSE
    # =====================================================

    course = Course.query.filter_by(

        id=course_id,

        instructor_id=current_user.id

    ).first_or_404()

    # =====================================================
    # ASSESSMENT
    # =====================================================

    assessment = Assessment.query.filter_by(

        id=assessment_id,

        course_id=course.id

    ).first_or_404()

    # =====================================================
    # QUESTION
    # =====================================================

    question = Question.query.filter_by(

        id=question_id,

        assessment_id=assessment.id

    ).first_or_404()

    # =====================================================
    # ENROLLMENT
    # =====================================================

    enrollment = Enrollment.query.filter_by(

        student_id=student_id,

        course_id=course.id

    ).first_or_404()

    # =====================================================
    # GET MARK
    # =====================================================

    mark_value = request.form.get(
        "mark",
        ""
    ).strip()

    if mark_value == "":

        flash(
            "Please enter a mark.",
            "danger"
        )

        return redirect(

            url_for(

                "assessment_student_marks",

                course_id=course.id,

                assessment_id=assessment.id

            )

        )

    # =====================================================
    # CONVERT
    # =====================================================

    try:

        mark = float(
            mark_value
        )

    except (
        TypeError,
        ValueError
    ):

        flash(
            "Invalid mark. Please enter a number.",
            "danger"
        )

        return redirect(

            url_for(

                "assessment_student_marks",

                course_id=course.id,

                assessment_id=assessment.id

            )

        )

    # =====================================================
    # MINIMUM
    # =====================================================

    if mark < 0:

        flash(
            "Student mark cannot be less than 0.",
            "danger"
        )

        return redirect(

            url_for(

                "assessment_student_marks",

                course_id=course.id,

                assessment_id=assessment.id

            )

        )

    # =====================================================
    # MAXIMUM
    # =====================================================

    if mark > question.max_mark:

        flash(

            f"Student mark cannot be greater than "
            f"{question.max_mark}.",

            "danger"

        )

        return redirect(

            url_for(

                "assessment_student_marks",

                course_id=course.id,

                assessment_id=assessment.id

            )

        )

    # =====================================================
    # FIND EXISTING RESULT
    # =====================================================

    result = (

        StudentQuestionResult.query

        .filter_by(

            enrollment_id=enrollment.id,

            question_id=question.id

        )

        .first()

    )

    # =====================================================
    # UPDATE
    # =====================================================

    if result:

        result.mark = mark

    # =====================================================
    # CREATE
    # =====================================================

    else:

        result = StudentQuestionResult(

            enrollment_id=enrollment.id,

            question_id=question.id,

            mark=mark

        )

        db.session.add(
            result
        )

    # =====================================================
    # SAVE
    # =====================================================

    try:

        db.session.commit()

    except Exception as error:

        db.session.rollback()

        print("====================================")
        print("UPDATE STUDENT QUESTION MARK ERROR")
        print("====================================")
        print(type(error).__name__)
        print(error)
        print("====================================")

        flash(
            "Unable to save student mark.",
            "danger"
        )

        return redirect(

            url_for(

                "assessment_student_marks",

                course_id=course.id,

                assessment_id=assessment.id

            )

        )

    # =====================================================
    # SUCCESS
    # =====================================================

    flash(

        f"Mark for Q{question.id} updated successfully.",

        "success"

    )

    return redirect(

        url_for(

            "assessment_student_marks",

            course_id=course.id,

            assessment_id=assessment.id

        )

    )


# =========================================================
# CLO ACHIEVEMENT SETTINGS
# =========================================================

CLO_PASS_PERCENTAGE = 50.0


# =========================================================
# HELPERS
# =========================================================

def get_course_enrollments(course_id):

    return (
        Enrollment.query
        .filter_by(course_id=course_id)
        .join(Student)
        .order_by(Student.student_number)
        .all()
    )


def get_assessment_questions(assessment_id):

    return (
        Question.query
        .filter_by(assessment_id=assessment_id)
        .order_by(Question.id)
        .all()
    )


def build_result_cache(enrollments, questions):

    enrollment_ids = [
        enrollment.id
        for enrollment in enrollments
    ]

    question_ids = [
        question.id
        for question in questions
    ]

    if not enrollment_ids or not question_ids:
        return {}

    results = (
        StudentQuestionResult.query
        .filter(
            StudentQuestionResult.enrollment_id.in_(enrollment_ids),
            StudentQuestionResult.question_id.in_(question_ids)
        )
        .all()
    )

    return {
        (
            result.enrollment_id,
            result.question_id
        ): float(result.mark or 0)
        for result in results
    }


# =========================================================
# STUDENT CLO ACHIEVEMENT
# =========================================================

def calculate_student_clo_achievement(
    enrollment,
    questions,
    clo,
    result_cache=None
):
    """
    Calculate one student's achievement for one CLO.

    If a question is mapped to multiple CLOs:
        question.max_mark / number_of_CLOs
        student_mark / number_of_CLOs

    Example:

        Question = 5 marks
        CLO1 + CLO2

        CLO1 available = 2.5
        CLO2 available = 2.5
    """

    available_marks = 0.0
    earned_marks = 0.0
    mapped_questions = 0

    for question in questions:

        if clo not in question.clos:
            continue

        clo_count = len(question.clos)

        if clo_count <= 0:
            continue

        mapped_questions += 1

        # -------------------------------------------------
        # Divide question maximum mark between CLOs
        # -------------------------------------------------

        allocated_max = (
            float(question.max_mark or 0)
            / clo_count
        )

        available_marks += allocated_max

        # -------------------------------------------------
        # Get student's mark
        # -------------------------------------------------

        if result_cache is not None:

            student_mark = result_cache.get(
                (
                    enrollment.id,
                    question.id
                ),
                0.0
            )

        else:

            result = (
                StudentQuestionResult.query
                .filter_by(
                    enrollment_id=enrollment.id,
                    question_id=question.id
                )
                .first()
            )

            student_mark = (
                float(result.mark or 0)
                if result
                else 0.0
            )

        # -------------------------------------------------
        # Divide student's mark between CLOs
        # -------------------------------------------------

        earned_marks += (
            float(student_mark)
            / clo_count
        )

    # =====================================================
    # NO AVAILABLE MARKS
    # =====================================================

    if available_marks <= 0:

        return {
            "available": 0.0,
            "earned": 0.0,
            "percentage": None,
            "status": "N/A",
            "mapped_questions": mapped_questions
        }

    # =====================================================
    # CALCULATE PERCENTAGE
    # =====================================================

    percentage = (
        earned_marks
        / available_marks
    ) * 100.0

    percentage = max(
        0.0,
        min(100.0, percentage)
    )

    return {
        "available": round(
            available_marks,
            4
        ),
        "earned": round(
            earned_marks,
            4
        ),
        "percentage": round(
            percentage,
            2
        ),
        "status": (
            "PASS"
            if percentage >= CLO_PASS_PERCENTAGE
            else "FAIL"
        ),
        "mapped_questions": mapped_questions
    }


# =========================================================
# ASSESSMENT CLO ACHIEVEMENT
# =========================================================

def calculate_assessment_clo_achievement(
    assessment,
    enrollments,
    questions,
    clo,
    result_cache=None
):
    """
    Calculate CLO achievement for one assessment.

    Formula:

        Total earned CLO marks
        -----------------------
        Total available CLO marks

    """

    total_available = 0.0
    total_earned = 0.0
    student_count = 0

    for enrollment in enrollments:

        result = calculate_student_clo_achievement(
            enrollment,
            questions,
            clo,
            result_cache=result_cache
        )

        if result["percentage"] is None:
            continue

        total_available += result["available"]
        total_earned += result["earned"]

        student_count += 1

    # =====================================================
    # NO DATA
    # =====================================================

    if total_available <= 0:

        return {
            "percentage": None,
            "status": "N/A",
            "student_count": 0,
            "available": 0.0,
            "earned": 0.0
        }

    # =====================================================
    # CALCULATE
    # =====================================================

    percentage = (
        total_earned
        / total_available
    ) * 100.0

    percentage = max(
        0.0,
        min(100.0, percentage)
    )

    return {
        "percentage": round(
            percentage,
            2
        ),
        "status": (
            "PASS"
            if percentage >= CLO_PASS_PERCENTAGE
            else "FAIL"
        ),
        "student_count": student_count,
        "available": round(
            total_available,
            4
        ),
        "earned": round(
            total_earned,
            4
        )
    }


# =========================================================
# STUDENT COURSE CLO ACHIEVEMENT
# =========================================================

def calculate_student_course_clo_achievement(
    enrollment,
    assessment_data,
    clo
):
    """
    Calculate one student's CLO achievement
    across the whole course.

    Assessment.weight is used as weighting factor.
    """

    weighted_sum = 0.0
    weight_sum = 0.0

    assessment_results = []

    for data in assessment_data:

        assessment = data["assessment"]
        questions = data["questions"]
        result_cache = data["result_cache"]

        result = calculate_student_clo_achievement(
            enrollment,
            questions,
            clo,
            result_cache=result_cache
        )

        if result["percentage"] is None:
            continue

        assessment_weight = float(
            assessment.weight or 0
        )

        assessment_results.append({
            "assessment": assessment,
            "percentage": result["percentage"],
            "earned": result["earned"],
            "available": result["available"],
            "weight": assessment_weight,
            "status": result["status"]
        })

        # -------------------------------------------------
        # Only positive assessment weights participate
        # -------------------------------------------------

        if assessment_weight > 0:

            weighted_sum += (
                result["percentage"]
                * assessment_weight
            )

            weight_sum += assessment_weight

    # =====================================================
    # NO WEIGHT
    # =====================================================

    if weight_sum <= 0:

        return {
            "percentage": None,
            "status": "N/A",
            "assessments": assessment_results
        }

    # =====================================================
    # FINAL CLO PERCENTAGE
    # =====================================================

    percentage = (
        weighted_sum
        / weight_sum
    )

    percentage = max(
        0.0,
        min(100.0, percentage)
    )

    return {
        "percentage": round(
            percentage,
            2
        ),
        "status": (
            "PASS"
            if percentage >= CLO_PASS_PERCENTAGE
            else "FAIL"
        ),
        "assessments": assessment_results
    }


# =========================================================
# COURSE CLO ACHIEVEMENT
# =========================================================

def calculate_course_clo_achievement(
    assessments,
    clo,
    assessment_data
):
    """
    Calculate one CLO at course level.

    Formula:

        Σ(Assessment CLO % × Assessment Weight)
        ----------------------------------------
              Σ(Assessment Weight)
    """

    weighted_sum = 0.0
    weight_sum = 0.0

    assessment_results = []

    for data in assessment_data:

        assessment = data["assessment"]
        questions = data["questions"]
        enrollments = data["enrollments"]
        result_cache = data["result_cache"]

        if not questions:
            continue

        result = calculate_assessment_clo_achievement(
            assessment,
            enrollments,
            questions,
            clo,
            result_cache=result_cache
        )

        if result["percentage"] is None:
            continue

        assessment_weight = float(
            assessment.weight or 0
        )

        assessment_results.append({
            "assessment": assessment,
            "percentage": result["percentage"],
            "status": result["status"],
            "student_count": result["student_count"],
            "weight": assessment_weight,
            "earned": result["earned"],
            "available": result["available"]
        })

        if assessment_weight > 0:

            weighted_sum += (
                result["percentage"]
                * assessment_weight
            )

            weight_sum += assessment_weight

    # =====================================================
    # NO DATA
    # =====================================================

    if weight_sum <= 0:

        return {
            "percentage": None,
            "status": "N/A",
            "assessments": assessment_results
        }

    percentage = (
        weighted_sum
        / weight_sum
    )

    percentage = max(
        0.0,
        min(100.0, percentage)
    )

    return {
        "percentage": round(
            percentage,
            2
        ),
        "status": (
            "PASS"
            if percentage >= CLO_PASS_PERCENTAGE
            else "FAIL"
        ),
        "assessments": assessment_results
    }


# =========================================================
# OVERALL CLO AVERAGE
# =========================================================

def calculate_overall_clo_average(clo_results):
    """
    Calculate overall average from valid CLO percentages.
    """

    valid_percentages = [
        item["percentage"]
        for item in clo_results
        if item["percentage"] is not None
    ]

    if not valid_percentages:
        return None

    return round(
        sum(valid_percentages)
        / len(valid_percentages),
        2
    )


# =========================================================
# STATUS HELPER
# =========================================================

def calculate_status(percentage):

    if percentage is None:
        return "N/A"

    return (
        "PASS"
        if percentage >= CLO_PASS_PERCENTAGE
        else "FAIL"
    )


# =========================================================
# PREPARE CLO DASHBOARD DATA
# =========================================================

def prepare_clo_dashboard_data(course):

    clos = (
        CLO.query
        .filter_by(course_id=course.id)
        .order_by(CLO.clo_code)
        .all()
    )

    assessments = (
        Assessment.query
        .filter_by(course_id=course.id)
        .order_by(Assessment.id)
        .all()
    )

    enrollments = get_course_enrollments(
        course.id
    )

    assessment_data = []

    for assessment in assessments:

        questions = get_assessment_questions(
            assessment.id
        )

        result_cache = build_result_cache(
            enrollments,
            questions
        )

        assessment_data.append({
            "assessment": assessment,
            "questions": questions,
            "enrollments": enrollments,
            "result_cache": result_cache
        })

    return (
        clos,
        assessments,
        enrollments,
        assessment_data
    )


# =========================================================
# BUILD COMPLETE COURSE REPORT DATA
# =========================================================

def build_course_clo_report_data(course):
    """
    Central function.

    Dashboard
    Report
    PDF

    all use the same calculations.
    """

    (
        clos,
        assessments,
        enrollments,
        assessment_data
    ) = prepare_clo_dashboard_data(course)

    # =====================================================
    # COURSE CLO
    # =====================================================

    course_clos = []

    for clo in clos:

        result = calculate_course_clo_achievement(
            assessments,
            clo,
            assessment_data
        )

        course_clos.append({
            "clo": clo,
            "percentage": result["percentage"],
            "status": result["status"],
            "assessments": result["assessments"]
        })

    # =====================================================
    # COURSE OVERALL
    # =====================================================

    course_average = (
        calculate_overall_clo_average(
            course_clos
        )
    )

    course_status = calculate_status(
        course_average
    )

    # =====================================================
    # ASSESSMENT CLO
    # =====================================================

    assessment_clos = []

    for data in assessment_data:

        assessment = data["assessment"]
        questions = data["questions"]
        result_cache = data["result_cache"]

        clo_results = []

        for clo in clos:

            result = calculate_assessment_clo_achievement(
                assessment,
                enrollments,
                questions,
                clo,
                result_cache=result_cache
            )

            clo_results.append({
                "clo": clo,
                "percentage": result["percentage"],
                "status": result["status"],
                "student_count": result["student_count"],
                "earned": result["earned"],
                "available": result["available"]
            })

        assessment_clos.append({
            "assessment": assessment,
            "clo_results": clo_results
        })

    # =====================================================
    # STUDENTS
    # =====================================================

    student_rows = []

    for enrollment in enrollments:

        clo_rows = []

        for clo in clos:

            result = calculate_student_course_clo_achievement(
                enrollment,
                assessment_data,
                clo
            )

            clo_rows.append({
                "clo": clo,
                "percentage": result["percentage"],
                "status": result["status"],
                "assessments": result["assessments"]
            })

        overall_percentage = (
            calculate_overall_clo_average(
                clo_rows
            )
        )

        overall_status = calculate_status(
            overall_percentage
        )

        student_rows.append({
            "enrollment": enrollment,
            "student": enrollment.student,
            "clos": clo_rows,
            "overall_percentage": overall_percentage,
            "overall_status": overall_status
        })

    # =====================================================
    # STATISTICS
    # =====================================================

    students_with_results = [
        row
        for row in student_rows
        if row["overall_percentage"] is not None
    ]

    passed_students = [
        row
        for row in students_with_results
        if row["overall_status"] == "PASS"
    ]

    failed_students = [
        row
        for row in students_with_results
        if row["overall_status"] == "FAIL"
    ]

    student_count = len(
        students_with_results
    )

    passed_count = len(
        passed_students
    )

    failed_count = len(
        failed_students
    )

    if student_count > 0:

        pass_rate = round(
            (
                passed_count
                / student_count
            ) * 100,
            2
        )

        fail_rate = round(
            (
                failed_count
                / student_count
            ) * 100,
            2
        )

    else:

        pass_rate = 0.0
        fail_rate = 0.0

    return {
        "clos": clos,
        "assessments": assessments,
        "enrollments": enrollments,
        "assessment_data": assessment_data,

        "course_clos": course_clos,
        "course_average": course_average,
        "course_status": course_status,

        "assessment_clos": assessment_clos,

        "student_rows": student_rows,

        "student_count": student_count,
        "passed_count": passed_count,
        "failed_count": failed_count,

        "pass_rate": pass_rate,
        "fail_rate": fail_rate
    }


# =========================================================
# ASSESSMENT CLO ACHIEVEMENT
# =========================================================

@app.route(
    "/courses/<int:course_id>/assessments/"
    "<int:assessment_id>/clo-achievement"
)
@login_required
def assessment_clo_achievement(
    course_id,
    assessment_id
):

    course = Course.query.filter_by(
        id=course_id,
        instructor_id=current_user.id
    ).first_or_404()

    assessment = Assessment.query.filter_by(
        id=assessment_id,
        course_id=course.id
    ).first_or_404()

    clos = (
        CLO.query
        .filter_by(course_id=course.id)
        .order_by(CLO.clo_code)
        .all()
    )

    enrollments = get_course_enrollments(
        course.id
    )

    questions = get_assessment_questions(
        assessment.id
    )

    result_cache = build_result_cache(
        enrollments,
        questions
    )

    # =====================================================
    # CLO RESULTS
    # =====================================================

    clo_results = []

    for clo in clos:

        result = calculate_assessment_clo_achievement(
            assessment,
            enrollments,
            questions,
            clo,
            result_cache=result_cache
        )

        clo_results.append({
            "clo": clo,
            "percentage": result["percentage"],
            "status": result["status"],
            "student_count": result["student_count"],
            "earned": result["earned"],
            "available": result["available"]
        })

    # =====================================================
    # STUDENT RESULTS
    # =====================================================

    student_results = []

    for enrollment in enrollments:

        student_clos = []

        for clo in clos:

            result = calculate_student_clo_achievement(
                enrollment,
                questions,
                clo,
                result_cache=result_cache
            )

            student_clos.append({
                "clo": clo,
                "percentage": result["percentage"],
                "earned": result["earned"],
                "available": result["available"],
                "status": result["status"]
            })

        overall_percentage = (
            calculate_overall_clo_average(
                student_clos
            )
        )

        overall_status = calculate_status(
            overall_percentage
        )

        student_results.append({
            "enrollment": enrollment,
            "student": enrollment.student,
            "clos": student_clos,
            "overall_percentage": overall_percentage,
            "overall_status": overall_status
        })

    return render_template(
        "assessment_clo_achievement.html",

        course=course,
        assessment=assessment,
        questions=questions,
        clos=clos,

        clo_results=clo_results,
        student_results=student_results,

        pass_percentage=CLO_PASS_PERCENTAGE
    )


# =========================================================
# COURSE CLO ACHIEVEMENT
# =========================================================

@app.route(
    "/courses/<int:course_id>/clo-achievement"
)
@login_required
def course_clo_achievement(course_id):

    course = Course.query.filter_by(
        id=course_id,
        instructor_id=current_user.id
    ).first_or_404()

    data = build_course_clo_report_data(
        course
    )

    return render_template(
        "course_clo_achievement.html",

        course=course,

        clos=data["clos"],
        assessments=data["assessments"],
        enrollments=data["enrollments"],

        clo_results=data["course_clos"],

        student_results=data["student_rows"],

        pass_percentage=CLO_PASS_PERCENTAGE
    )


# =========================================================
# COURSE CLO ACHIEVEMENT REPORT
# =========================================================

@app.route(
    "/courses/<int:course_id>/clo-achievement/report"
)
@login_required
def course_clo_achievement_report(course_id):

    course = Course.query.filter_by(
        id=course_id,
        instructor_id=current_user.id
    ).first_or_404()

    data = build_course_clo_report_data(
        course
    )

    return render_template(
        "course_clo_achievement_report.html",

        course=course,

        clos=data["clos"],
        assessments=data["assessments"],
        enrollments=data["enrollments"],

        course_clos=data["course_clos"],

        course_average=data["course_average"],
        course_status=data["course_status"],

        assessment_clos=data["assessment_clos"],

        student_rows=data["student_rows"],

        student_count=data["student_count"],
        passed_count=data["passed_count"],
        failed_count=data["failed_count"],

        pass_rate=data["pass_rate"],
        fail_rate=data["fail_rate"],

        pass_percentage=CLO_PASS_PERCENTAGE
    )


# =========================================================
# COURSE DASHBOARD
# =========================================================

@app.route(
    "/courses/<int:course_id>/dashboard"
)
@login_required
def course_dashboard(course_id):

    course = Course.query.filter_by(
        id=course_id,
        instructor_id=current_user.id
    ).first_or_404()

    data = build_course_clo_report_data(
        course
    )

    return render_template(
        "course_dashboard.html",

        course=course,

        clos=data["clos"],
        assessments=data["assessments"],
        enrollments=data["enrollments"],

        course_clos=data["course_clos"],

        assessment_clos=data["assessment_clos"],

        student_rows=data["student_rows"],

        course_average=data["course_average"],

        course_status=data["course_status"],

        student_count=data["student_count"],
        passed_count=data["passed_count"],
        failed_count=data["failed_count"],

        pass_rate=data["pass_rate"],
        fail_rate=data["fail_rate"],

        pass_percentage=CLO_PASS_PERCENTAGE
    )


# =========================================================
# STUDENT CLO ACHIEVEMENT
# =========================================================

@app.route(
    "/courses/<int:course_id>/students/"
    "<int:student_id>/clo-achievement"
)
@login_required
def student_clo_achievement(
    course_id,
    student_id
):

    course = Course.query.filter_by(
        id=course_id,
        instructor_id=current_user.id
    ).first_or_404()

    enrollment = Enrollment.query.filter_by(
        course_id=course.id,
        student_id=student_id
    ).first_or_404()

    student = enrollment.student

    data = build_course_clo_report_data(
        course
    )

    # =====================================================
    # FIND STUDENT
    # =====================================================

    student_row = next(
        (
            row
            for row in data["student_rows"]
            if row["enrollment"].id == enrollment.id
        ),
        None
    )

    if student_row is None:

        clo_results = []

        overall_percentage = None
        overall_status = "N/A"

    else:

        clo_results = student_row["clos"]

        overall_percentage = (
            student_row["overall_percentage"]
        )

        overall_status = (
            student_row["overall_status"]
        )

    return render_template(
        "student_clo_achievement.html",

        course=course,

        student=student,
        enrollment=enrollment,

        clos=data["clos"],
        assessments=data["assessments"],

        clo_results=clo_results,

        overall_percentage=overall_percentage,
        overall_status=overall_status,

        pass_percentage=CLO_PASS_PERCENTAGE
    )



# =========================================================
# PDF CHART - COURSE CLO ACHIEVEMENT
# =========================================================

def create_course_clo_chart(course_clos):

    drawing = Drawing(
        500,
        280
    )

    chart = VerticalBarChart()

    chart.x = 55
    chart.y = 55

    chart.height = 190
    chart.width = 410

    values = []
    labels = []

    for item in course_clos:

        clo = item.get("clo")

        labels.append(
            str(
                getattr(
                    clo,
                    "clo_code",
                    ""
                )
            )
        )

        percentage_value = item.get(
            "percentage",
            0
        )

        try:
            percentage_value = float(
                percentage_value
                if percentage_value is not None
                else 0
            )
        except (
            TypeError,
            ValueError
        ):
            percentage_value = 0

        percentage_value = max(
            0,
            min(
                percentage_value,
                100
            )
        )

        values.append(
            percentage_value
        )

    if not values:

        drawing.add(
            String(
                250,
                130,
                "No CLO achievement data available",
                textAnchor="middle",
                fontName="Helvetica",
                fontSize=10
            )
        )

        return drawing

    chart.data = [
        values
    ]

    chart.categoryAxis.categoryNames = labels

    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = 100
    chart.valueAxis.valueStep = 10

    chart.categoryAxis.labels.fontName = (
        "Helvetica-Bold"
    )

    chart.categoryAxis.labels.fontSize = 8

    chart.categoryAxis.labels.angle = 0

    chart.valueAxis.labels.fontName = (
        "Helvetica"
    )

    chart.valueAxis.labels.fontSize = 7

    chart.barWidth = 25
    chart.groupSpacing = 15

    chart.barLabels.nudge = 7
    chart.barLabels.fontName = (
        "Helvetica-Bold"
    )
    chart.barLabels.fontSize = 8
    chart.barLabels.boxAnchor = "s"
    chart.barLabels.dy = 5

    chart.barLabelFormat = "%.2f%%"

    chart.bars[0].fillColor = (
        colors.HexColor("#12355B")
    )

    chart.bars[0].strokeColor = (
        colors.HexColor("#12355B")
    )

    chart.valueAxis.gridStrokeColor = (
        colors.HexColor("#D9E1E8")
    )

    chart.valueAxis.gridStrokeWidth = 0.5

    chart.categoryAxis.strokeColor = (
        colors.HexColor("#777777")
    )

    chart.valueAxis.strokeColor = (
        colors.HexColor("#777777")
    )

    drawing.add(chart)

    return drawing


# =========================================================
# PDF CHART - ASSESSMENT CLO ACHIEVEMENT
# =========================================================

def create_assessment_clo_chart(
    assessment_item
):

    drawing = Drawing(
        500,
        280
    )

    chart = VerticalBarChart()

    chart.x = 55
    chart.y = 55

    chart.height = 190
    chart.width = 410

    values = []
    labels = []

    clo_results = assessment_item.get(
        "clo_results",
        []
    )

    for item in clo_results:

        clo = item.get("clo")

        labels.append(
            str(
                getattr(
                    clo,
                    "clo_code",
                    ""
                )
            )
        )

        percentage_value = item.get(
            "percentage",
            0
        )

        try:
            percentage_value = float(
                percentage_value
                if percentage_value is not None
                else 0
            )
        except (
            TypeError,
            ValueError
        ):
            percentage_value = 0

        percentage_value = max(
            0,
            min(
                percentage_value,
                100
            )
        )

        values.append(
            percentage_value
        )

    if not values:

        drawing.add(
            String(
                250,
                130,
                "No assessment CLO data available",
                textAnchor="middle",
                fontName="Helvetica",
                fontSize=10
            )
        )

        return drawing

    chart.data = [
        values
    ]

    chart.categoryAxis.categoryNames = labels

    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = 100
    chart.valueAxis.valueStep = 10

    chart.categoryAxis.labels.fontName = (
        "Helvetica-Bold"
    )

    chart.categoryAxis.labels.fontSize = 8

    chart.valueAxis.labels.fontName = (
        "Helvetica"
    )

    chart.valueAxis.labels.fontSize = 7

    chart.barWidth = 25
    chart.groupSpacing = 15

    chart.barLabels.nudge = 7
    chart.barLabels.fontName = (
        "Helvetica-Bold"
    )
    chart.barLabels.fontSize = 8
    chart.barLabels.boxAnchor = "s"
    chart.barLabels.dy = 5

    chart.barLabelFormat = "%.2f%%"

    chart.bars[0].fillColor = (
        colors.HexColor("#2E7D32")
    )

    chart.bars[0].strokeColor = (
        colors.HexColor("#2E7D32")
    )

    chart.valueAxis.gridStrokeColor = (
        colors.HexColor("#D9E1E8")
    )

    chart.valueAxis.gridStrokeWidth = 0.5

    chart.categoryAxis.strokeColor = (
        colors.HexColor("#777777")
    )

    chart.valueAxis.strokeColor = (
        colors.HexColor("#777777")
    )

    drawing.add(chart)

    return drawing


# =========================================================
# PDF CHART - PASS / FAIL
# =========================================================

def create_pass_fail_chart(
    passed_count,
    failed_count
):

    drawing = Drawing(
        500,
        300
    )

    try:
        passed = float(
            passed_count or 0
        )
    except (
        TypeError,
        ValueError
    ):
        passed = 0

    try:
        failed = float(
            failed_count or 0
        )
    except (
        TypeError,
        ValueError
    ):
        failed = 0

    passed = max(
        0,
        passed
    )

    failed = max(
        0,
        failed
    )

    values = [
        passed,
        failed
    ]

    labels = [
        "Passed",
        "Failed"
    ]

    # -----------------------------------------------------
    # If there are no students
    # -----------------------------------------------------

    if passed == 0 and failed == 0:

        drawing.add(
            String(
                250,
                145,
                "No student pass/fail data available",
                textAnchor="middle",
                fontName="Helvetica",
                fontSize=10
            )
        )

        return drawing

    chart = VerticalBarChart()

    chart.x = 70
    chart.y = 60

    chart.height = 190
    chart.width = 360

    # Two series so Passed and Failed have
    # different colors.

    chart.data = [
        [
            passed,
            0
        ],
        [
            0,
            failed
        ]
    ]

    chart.categoryAxis.categoryNames = labels

    max_value = max(
        values
    )

    if max_value <= 0:
        max_value = 1

    chart.valueAxis.valueMin = 0

    chart.valueAxis.valueMax = (
        max_value
        + max(
            1,
            max_value * 0.20
        )
    )

    chart.valueAxis.valueStep = max(
        1,
        int(
            max_value / 5
        )
    )

    chart.categoryAxis.labels.fontName = (
        "Helvetica-Bold"
    )

    chart.categoryAxis.labels.fontSize = 10

    chart.valueAxis.labels.fontName = (
        "Helvetica"
    )

    chart.valueAxis.labels.fontSize = 8

    chart.barWidth = 45
    chart.groupSpacing = 25

    # -----------------------------------------------------
    # Passed bar
    # -----------------------------------------------------

    chart.bars[0].fillColor = (
        colors.HexColor("#2E7D32")
    )

    chart.bars[0].strokeColor = (
        colors.HexColor("#2E7D32")
    )

    # -----------------------------------------------------
    # Failed bar
    # -----------------------------------------------------

    chart.bars[1].fillColor = (
        colors.HexColor("#C62828")
    )

    chart.bars[1].strokeColor = (
        colors.HexColor("#C62828")
    )

    chart.valueAxis.gridStrokeColor = (
        colors.HexColor("#D9E1E8")
    )

    chart.valueAxis.gridStrokeWidth = 0.5

    chart.categoryAxis.strokeColor = (
        colors.HexColor("#777777")
    )

    chart.valueAxis.strokeColor = (
        colors.HexColor("#777777")
    )

    # -----------------------------------------------------
    # Do not use automatic labels here because there are
    # two series and one series contains zero values.
    # Add values manually.
    # -----------------------------------------------------

    drawing.add(
        chart
    )

    drawing.add(
        String(
            160,
            255,
            f"Passed: {int(passed)}",
            textAnchor="middle",
            fontName="Helvetica-Bold",
            fontSize=9,
            fillColor=colors.HexColor("#2E7D32")
        )
    )

    drawing.add(
        String(
            340,
            255,
            f"Failed: {int(failed)}",
            textAnchor="middle",
            fontName="Helvetica-Bold",
            fontSize=9,
            fillColor=colors.HexColor("#C62828")
        )
    )

    return drawing


# =========================================================
# GENERATE OFFICIAL COURSE CLO PDF
# =========================================================

def generate_course_clo_pdf(
    course,
    clos,
    assessments,
    enrollments,
    course_clos,
    course_average,
    course_status,
    assessment_clos,
    student_rows,
    student_count,
    passed_count,
    failed_count,
    pass_rate,
    fail_rate,
    pass_percentage
):

    buffer = BytesIO()

    # =====================================================
    # PAGE SETUP
    # =====================================================

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,

        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,

        title="Course CLO Achievement Report",
        author="CLO Achievement Management System"
    )

    styles = getSampleStyleSheet()

    # =====================================================
    # STYLES
    # =====================================================

    university_style = ParagraphStyle(
        "University",
        parent=styles["Title"],
        fontSize=20,
        leading=24,
        alignment=TA_CENTER,
        spaceAfter=8,
        textColor=colors.HexColor("#12355B")
    )

    college_style = ParagraphStyle(
        "College",
        parent=styles["Normal"],
        fontSize=12,
        leading=16,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#444444")
    )

    report_title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontSize=22,
        leading=28,
        alignment=TA_CENTER,
        spaceBefore=35,
        spaceAfter=20,
        textColor=colors.HexColor("#12355B")
    )

    section_style = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontSize=14,
        leading=18,
        spaceBefore=12,
        spaceAfter=10,
        textColor=colors.HexColor("#12355B")
    )

    chart_title_style = ParagraphStyle(
        "ChartTitle",
        parent=styles["Heading3"],
        fontSize=11,
        leading=14,
        alignment=TA_CENTER,
        spaceBefore=6,
        spaceAfter=5,
        textColor=colors.HexColor("#12355B")
    )

    normal_style = ParagraphStyle(
        "NormalCustom",
        parent=styles["Normal"],
        fontSize=9,
        leading=13
    )

    small_style = ParagraphStyle(
        "Small",
        parent=styles["Normal"],
        fontSize=8,
        leading=10
    )

    center_style = ParagraphStyle(
        "Center",
        parent=normal_style,
        alignment=TA_CENTER
    )

    student_cell_style = ParagraphStyle(
        "StudentCell",
        parent=styles["Normal"],
        fontSize=6,
        leading=7,
        alignment=TA_CENTER
    )

    student_name_style = ParagraphStyle(
        "StudentName",
        parent=styles["Normal"],
        fontSize=6,
        leading=7,
        alignment=TA_LEFT
    )

    student_header_style = ParagraphStyle(
        "StudentHeader",
        parent=styles["Normal"],
        fontSize=6,
        leading=7,
        alignment=TA_CENTER,
        textColor=colors.white
    )

    # =====================================================
    # HELPERS
    # =====================================================

    def safe_text(value):

        if value is None:
            return ""

        return str(value)

    def percentage(value):

        if value is None:
            return "N/A"

        try:

            return (
                f"{float(value):.2f}%"
            )

        except (
            TypeError,
            ValueError
        ):

            return "N/A"

    def status_text(status):

        return safe_text(
            status
        )

    def make_table(
        table_data,
        col_widths=None,
        header=True,
        font_size=8
    ):

        table = Table(
            table_data,
            colWidths=col_widths,
            repeatRows=1 if header else 0
        )

        commands = [

            (
                "FONTNAME",
                (0, 0),
                (-1, -1),
                "Helvetica"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                font_size
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#B8C2CC")
            ),

            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                4
            ),

            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                4
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                4
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                4
            )
        ]

        if header:

            commands.extend([

                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#12355B")
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),

                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),

                (
                    "ALIGN",
                    (0, 0),
                    (-1, 0),
                    "CENTER"
                )
            ])

        table.setStyle(
            TableStyle(
                commands
            )
        )

        return table

    # =====================================================
    # STORY
    # =====================================================

    story = []

    # =====================================================
    # COVER PAGE
    # =====================================================

    story.append(
        Spacer(
            1,
            25 * mm
        )
    )

    story.append(
        Paragraph(
            "UNIVERSITY NAME",
            university_style
        )
    )

    story.append(
        Paragraph(
            "COLLEGE OF ______________________________",
            college_style
        )
    )

    story.append(
        Paragraph(
            "DEPARTMENT OF ____________________________",
            college_style
        )
    )

    story.append(
        Spacer(
            1,
            20 * mm
        )
    )

    story.append(
        Paragraph(
            "COURSE CLO ACHIEVEMENT REPORT",
            report_title_style
        )
    )

    story.append(
        Paragraph(
            "Official Academic Assessment Report",
            college_style
        )
    )

    story.append(
        Spacer(
            1,
            20 * mm
        )
    )

    # =====================================================
    # COURSE INFORMATION
    # =====================================================

    course_code = safe_text(
        getattr(
            course,
            "course_code",
            ""
        )
    )

    course_name = safe_text(
        getattr(
            course,
            "course_name",
            ""
        )
    )

    academic_info = [

        [
            "Course Code",
            course_code
        ],

        [
            "Course Name",
            course_name
        ],

        [
            "Number of CLOs",
            str(
                len(clos)
            )
        ],

        [
            "Number of Assessments",
            str(
                len(assessments)
            )
        ],

        [
            "Number of Students",
            str(
                len(enrollments)
            )
        ],

        [
            "Report Date",
            datetime.now().strftime(
                "%d/%m/%Y"
            )
        ]
    ]

    course_info_table = Table(
        academic_info,
        colWidths=[
            55 * mm,
            105 * mm
        ]
    )

    course_info_table.setStyle(
        TableStyle([

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#B8C2CC")
            ),

            (
                "BACKGROUND",
                (0, 0),
                (0, -1),
                colors.HexColor("#EAF0F6")
            ),

            (
                "FONTNAME",
                (0, 0),
                (0, -1),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                10
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                8
            )
        ])
    )

    story.append(
        course_info_table
    )

    story.append(
        Spacer(
            1,
            25 * mm
        )
    )

    story.append(
        Paragraph(
            "Prepared by: Course Instructor",
            center_style
        )
    )

    story.append(
        Spacer(
            1,
            8 * mm
        )
    )

    story.append(
        Paragraph(
            "Official Academic Document",
            center_style
        )
    )

    story.append(
        PageBreak()
    )

    # =====================================================
    # 1. EXECUTIVE SUMMARY
    # =====================================================

    story.append(
        Paragraph(
            "1. Executive Summary",
            section_style
        )
    )

    summary_data = [

        [
            "Indicator",
            "Value"
        ],

        [
            "Overall Course CLO Achievement",
            percentage(
                course_average
            )
        ],

        [
            "Course Status",
            status_text(
                course_status
            )
        ],

        [
            "CLO Pass Threshold",
            percentage(
                pass_percentage
            )
        ],

        [
            "Students with Results",
            str(
                student_count
            )
        ],

        [
            "Passed Students",
            str(
                passed_count
            )
        ],

        [
            "Failed Students",
            str(
                failed_count
            )
        ],

        [
            "Pass Rate",
            percentage(
                pass_rate
            )
        ],

        [
            "Fail Rate",
            percentage(
                fail_rate
            )
        ]
    ]

    story.append(
        make_table(
            summary_data,
            col_widths=[
                100 * mm,
                55 * mm
            ]
        )
    )

    # =====================================================
    # PASS / FAIL CHART
    # =====================================================

    story.append(
        Spacer(
            1,
            8 * mm
        )
    )

    story.append(
        Paragraph(
            "Student Pass / Fail Distribution",
            chart_title_style
        )
    )

    story.append(
        create_pass_fail_chart(
            passed_count,
            failed_count
        )
    )

    # =====================================================
    # 2. COURSE CLO ACHIEVEMENT
    # =====================================================

    story.append(
        PageBreak()
    )

    story.append(
        Paragraph(
            "2. Course CLO Achievement",
            section_style
        )
    )

    course_clo_data = [

        [
            "CLO",
            "Description",
            "Achievement",
            "Status"
        ]
    ]

    for item in course_clos:

        clo = item.get(
            "clo"
        )

        description = safe_text(
            getattr(
                clo,
                "description",
                ""
            )
        )

        course_clo_data.append([

            safe_text(
                getattr(
                    clo,
                    "clo_code",
                    ""
                )
            ),

            Paragraph(
                description,
                small_style
            ),

            percentage(
                item.get(
                    "percentage"
                )
            ),

            status_text(
                item.get(
                    "status"
                )
            )
        ])

    story.append(
        make_table(
            course_clo_data,
            col_widths=[
                22 * mm,
                85 * mm,
                30 * mm,
                25 * mm
            ]
        )
    )

    # =====================================================
    # COURSE CLO BAR CHART
    # =====================================================

    story.append(
        Spacer(
            1,
            8 * mm
        )
    )

    story.append(
        Paragraph(
            "Course CLO Achievement Bar Chart",
            chart_title_style
        )
    )

    story.append(
        create_course_clo_chart(
            course_clos
        )
    )

    # =====================================================
    # 3. ASSESSMENT CLO ACHIEVEMENT
    # =====================================================

    story.append(
        PageBreak()
    )

    story.append(
        Paragraph(
            "3. Assessment CLO Achievement",
            section_style
        )
    )

    for assessment_index, assessment_item in enumerate(
        assessment_clos,
        start=1
    ):

        assessment = assessment_item.get(
            "assessment"
        )

        assessment_name = safe_text(
            getattr(
                assessment,
                "name",
                ""
            )
        )

        assessment_type = safe_text(
            getattr(
                assessment,
                "assessment_type",
                ""
            )
        )

        assessment_weight = float(
            getattr(
                assessment,
                "weight",
                0
            ) or 0
        )

        story.append(
            Paragraph(
                f"<b>{assessment_name}</b> "
                f"({assessment_type}) - "
                f"Weight: "
                f"{assessment_weight:.2f}%",
                normal_style
            )
        )

        story.append(
            Spacer(
                1,
                3 * mm
            )
        )

        assessment_table = [

            [
                "CLO",
                "Achievement",
                "Status",
                "Students"
            ]
        ]

        for item in assessment_item.get(
            "clo_results",
            []
        ):

            assessment_table.append([

                safe_text(
                    getattr(
                        item.get("clo"),
                        "clo_code",
                        ""
                    )
                ),

                percentage(
                    item.get(
                        "percentage"
                    )
                ),

                status_text(
                    item.get(
                        "status"
                    )
                ),

                str(
                    item.get(
                        "student_count",
                        0
                    )
                )
            ])

        story.append(
            make_table(
                assessment_table,
                col_widths=[
                    35 * mm,
                    45 * mm,
                    40 * mm,
                    35 * mm
                ]
            )
        )

        # =================================================
        # ASSESSMENT CLO BAR CHART
        # =================================================

        story.append(
            Spacer(
                1,
                5 * mm
            )
        )

        story.append(
            Paragraph(
                f"{assessment_name} - CLO Achievement",
                chart_title_style
            )
        )

        story.append(
            create_assessment_clo_chart(
                assessment_item
            )
        )

        # -------------------------------------------------
        # Page break between assessments
        # -------------------------------------------------

        if (
            assessment_index
            < len(assessment_clos)
        ):

            story.append(
                PageBreak()
            )

    # =====================================================
    # 4. STUDENT CLO ACHIEVEMENT
    # =====================================================

    story.append(
        PageBreak()
    )

    story.append(
        Paragraph(
            "4. Student CLO Achievement",
            section_style
        )
    )

    student_header = [

        "No.",
        "Student ID",
        "Student Name"
    ]

    for clo in clos:

        student_header.append(
            safe_text(
                getattr(
                    clo,
                    "clo_code",
                    ""
                )
            )
        )

    student_header.extend([
        "Overall",
        "Status"
    ])

    student_table = [

        [
            Paragraph(
                safe_text(
                    value
                ),
                student_header_style
            )

            for value
            in student_header
        ]
    ]

    for index, row in enumerate(
        student_rows,
        start=1
    ):

        student = row.get(
            "student"
        )

        student_number = safe_text(
            getattr(
                student,
                "student_number",
                ""
            )
        )

        fname = safe_text(
            getattr(
                student,
                "fname",
                ""
            )
        )

        lname = safe_text(
            getattr(
                student,
                "lname",
                ""
            )
        )

        student_name = (
            f"{fname} {lname}"
        ).strip()

        student_row = [

            Paragraph(
                str(index),
                student_cell_style
            ),

            Paragraph(
                student_number,
                student_cell_style
            ),

            Paragraph(
                student_name,
                student_name_style
            )
        ]

        for clo_item in row.get(
            "clos",
            []
        ):

            student_row.append(
                Paragraph(
                    percentage(
                        clo_item.get(
                            "percentage"
                        )
                    ),
                    student_cell_style
                )
            )

        student_row.extend([

            Paragraph(
                percentage(
                    row.get(
                        "overall_percentage"
                    )
                ),
                student_cell_style
            ),

            Paragraph(
                status_text(
                    row.get(
                        "overall_status"
                    )
                ),
                student_cell_style
            )
        ])

        student_table.append(
            student_row
        )

    # =====================================================
    # STUDENT TABLE WIDTH
    # =====================================================

    available_width = (
        A4[0]
        - doc.leftMargin
        - doc.rightMargin
    )

    no_width = 9 * mm
    id_width = 25 * mm
    name_width = 45 * mm
    overall_width = 22 * mm
    status_width = 20 * mm

    clo_count = len(
        clos
    )

    fixed_width = (
        no_width
        + id_width
        + name_width
        + overall_width
        + status_width
    )

    remaining_width = (
        available_width
        - fixed_width
    )

    if clo_count > 0:

        clo_width = (
            remaining_width
            / clo_count
        )

    else:

        clo_width = 15 * mm

    clo_width = max(
        clo_width,
        10 * mm
    )

    clo_width = min(
        clo_width,
        18 * mm
    )

    student_widths = (

        [
            no_width,
            id_width,
            name_width
        ]

        +

        [
            clo_width
            for _ in range(
                clo_count
            )
        ]

        +

        [
            overall_width,
            status_width
        ]
    )

    total_width = sum(
        student_widths
    )

    if total_width > available_width:

        scale = (
            available_width
            / total_width
        )

        student_widths = [

            width * scale

            for width
            in student_widths
        ]

    total_width = sum(
        student_widths
    )

    if (
        total_width
        < available_width
        and clo_count > 0
    ):

        difference = (
            available_width
            - total_width
        )

        extra_per_clo = (
            difference
            / clo_count
        )

        for i in range(
            3,
            3 + clo_count
        ):

            student_widths[i] += (
                extra_per_clo
            )

    # =====================================================
    # STUDENT PDF TABLE
    # =====================================================

    student_pdf_table = Table(
        student_table,
        colWidths=student_widths,
        repeatRows=1,
        splitByRow=1
    )

    student_pdf_table.setStyle(
        TableStyle([

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.4,
                colors.HexColor("#B8C2CC")
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#12355B")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, 0),
                6
            ),

            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                2
            ),

            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                2
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                3
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                3
            ),

            (
                "ALIGN",
                (2, 1),
                (2, -1),
                "LEFT"
            )
        ])
    )

    story.append(
        student_pdf_table
    )

    # =====================================================
    # 5. STATISTICAL SUMMARY
    # =====================================================

    story.append(
        PageBreak()
    )

    story.append(
        Paragraph(
            "5. Statistical Summary",
            section_style
        )
    )

    statistics_data = [

        [
            "Statistic",
            "Value"
        ],

        [
            "Students with Results",
            str(
                student_count
            )
        ],

        [
            "Passed Students",
            str(
                passed_count
            )
        ],

        [
            "Failed Students",
            str(
                failed_count
            )
        ],

        [
            "Pass Rate",
            percentage(
                pass_rate
            )
        ],

        [
            "Fail Rate",
            percentage(
                fail_rate
            )
        ],

        [
            "Overall Course Achievement",
            percentage(
                course_average
            )
        ]
    ]

    story.append(
        make_table(
            statistics_data,
            col_widths=[
                100 * mm,
                55 * mm
            ]
        )
    )

    # =====================================================
    # PASS / FAIL CHART - STATISTICAL SECTION
    # =====================================================

    story.append(
        Spacer(
            1,
            8 * mm
        )
    )

    story.append(
        Paragraph(
            "Pass / Fail Bar Chart",
            chart_title_style
        )
    )

    story.append(
        create_pass_fail_chart(
            passed_count,
            failed_count
        )
    )

    story.append(
        Spacer(
            1,
            10 * mm
        )
    )

    # =====================================================
    # 6. APPROVAL
    # =====================================================

    story.append(
        Paragraph(
            "6. Review and Approval",
            section_style
        )
    )

    approval_data = [

        [
            "Prepared By",
            "Reviewed By",
            "Approved By"
        ],

        [
            "Course Instructor",
            "Program Coordinator",
            "Head of Department"
        ],

        [
            "\n\nSignature: ____________________\n\n"
            "Date: ________________________",

            "\n\nSignature: ____________________\n\n"
            "Date: ________________________",

            "\n\nSignature: ____________________\n\n"
            "Date: ________________________"
        ]
    ]

    approval_table = Table(
        approval_data,
        colWidths=[
            55 * mm,
            55 * mm,
            55 * mm
        ]
    )

    approval_table.setStyle(
        TableStyle([

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#B8C2CC")
            ),

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#12355B")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 1),
                "Helvetica-Bold"
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                8
            )
        ])
    )

    story.append(
        approval_table
    )

    story.append(
        Spacer(
            1,
            15 * mm
        )
    )

    story.append(
        Paragraph(
            "This document is an official academic report "
            "generated by the CLO Achievement Management System.",
            center_style
        )
    )

    # =====================================================
    # FOOTER
    # =====================================================

    def add_page_number(
        canvas,
        doc
    ):

        canvas.saveState()

        width, height = A4

        canvas.setStrokeColor(
            colors.HexColor("#B8C2CC")
        )

        canvas.line(
            15 * mm,
            12 * mm,
            width - 15 * mm,
            12 * mm
        )

        canvas.setFont(
            "Helvetica",
            7
        )

        canvas.setFillColor(
            colors.HexColor("#666666")
        )

        canvas.drawString(
            15 * mm,
            7 * mm,
            "Official Academic Report"
        )

        canvas.drawRightString(
            width - 15 * mm,
            7 * mm,
            f"Page {doc.page}"
        )

        canvas.restoreState()

    # =====================================================
    # BUILD PDF
    # =====================================================

    doc.build(
        story,
        onFirstPage=add_page_number,
        onLaterPages=add_page_number
    )

    buffer.seek(0)

    return buffer


# =========================================================
# DOWNLOAD OFFICIAL COURSE CLO PDF
# =========================================================

@app.route(
    "/courses/<int:course_id>/clo-achievement/report/pdf"
)
@login_required
def download_course_clo_pdf(
    course_id
):

    # =====================================================
    # GET COURSE
    # =====================================================

    course = Course.query.filter_by(
        id=course_id,
        instructor_id=current_user.id
    ).first_or_404()

    # =====================================================
    # BUILD ALL REPORT DATA
    # =====================================================

    data = build_course_clo_report_data(
        course
    )

    # =====================================================
    # GENERATE PDF
    # =====================================================

    pdf_buffer = generate_course_clo_pdf(

        course=course,

        clos=data["clos"],

        assessments=data["assessments"],

        enrollments=data["enrollments"],

        course_clos=data["course_clos"],

        course_average=data["course_average"],

        course_status=data["course_status"],

        assessment_clos=data["assessment_clos"],

        student_rows=data["student_rows"],

        student_count=data["student_count"],

        passed_count=data["passed_count"],

        failed_count=data["failed_count"],

        pass_rate=data["pass_rate"],

        fail_rate=data["fail_rate"],

        pass_percentage=CLO_PASS_PERCENTAGE
    )

    # =====================================================
    # FILE NAME
    # =====================================================

    course_code = (
        str(
            course.course_code
        )
        .strip()
        .replace(
            " ",
            "_"
        )
        .replace(
            "/",
            "-"
        )
    )

    filename = (
        f"{course_code}_"
        f"CLO_Achievement_Report_"
        f"{datetime.now().strftime('%Y-%m-%d')}"
        f".pdf"
    )

    # =====================================================
    # DOWNLOAD
    # =====================================================

    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )